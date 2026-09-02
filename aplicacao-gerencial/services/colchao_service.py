from __future__ import annotations

import copy
import calendar
import csv
import io
import json
import shutil
import threading
import unicodedata
import zipfile
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries


class ColchaoError(RuntimeError):
    pass


DEFAULT_CONFIG = {
    "excel_path": "",
    "main_sheet": "",
}

PROFILE_DEFAULTS = {
    "alpha": {
        "name": "Alpha",
        "excel_path": "",
        "main_sheet": "",
        "sheet_options": [],
        "id_labels": ["DEBIT ID"],
        "client_labels": ["CLIENTE"],
        "money_labels": ["VALOR DO ACORDO"],
        "due_labels": ["DATA DO VENCIMENTO"],
        "operator_labels": ["OPERADOR"],
        "installment_labels": ["PARCELAS"],
        "required_headers": [],
    },
    "beta": {
        "name": "Beta",
        "excel_path": str(Path.home() / "Downloads" / "COLCHOES" / "COLCHÃO TOTAL BETA.xlsx"),
        "main_sheet": "ATIVO",
        "sheet_options": ["ATIVO", "PASSIVO"],
        "id_labels": ["SUITID", "SUIT"],
        "client_labels": ["NOME", "CLIENTE"],
        "money_labels": ["CASH", "VALOR DO ACORDO"],
        "due_labels": ["MÊS", "MES", "DATA DO VENCIMENTO", "MÊS DE EXPIRAÇÃO", "MES DE EXPIRACAO"],
        "operator_labels": ["OPERADOR", "OPERADORES"],
        "installment_labels": ["COND PARCELADAS", "COND PARCELAS", "PARCELAS"],
        "required_headers": ["STATUS", "ACORDO"],
    },
}

GENERIC_PROFILE_DEFAULT = {
    "name": "Colchao",
    "excel_path": "",
    "main_sheet": "COLCHAO",
    "sheet_options": ["COLCHAO"],
    "id_labels": ["IDENTIFICADOR", "NPJ", "DEBIT ID", "SUITID"],
    "client_labels": ["CLIENTE", "NOME"],
    "money_labels": ["VALOR DO ACORDO", "VALOR", "CASH"],
    "due_labels": ["DATA DO VENCIMENTO", "VENCIMENTO", "MES"],
    "operator_labels": ["OPERADOR", "NEGOCIADOR"],
    "installment_labels": ["PARCELAS", "PARCELA"],
    "required_headers": [],
}

DEFAULT_COLCHAO_FIELDS = [
    {"key": "identifier", "label": "Identificador", "type": "text", "required": True, "enabled": True, "role": "identifier"},
    {"key": "client", "label": "Cliente", "type": "text", "required": True, "enabled": True, "role": "client"},
    {"key": "document", "label": "CPF/CNPJ", "type": "text", "required": False, "enabled": True, "role": "document"},
    {"key": "process", "label": "Processo", "type": "text", "required": False, "enabled": False, "role": "process"},
    {"key": "agreement_type", "label": "Tipo de acordo", "type": "select", "required": True, "enabled": True, "role": "agreement_type", "options": ["A VISTA", "PARCELADO"]},
    {"key": "total_value", "label": "Valor total do acordo", "type": "money", "required": True, "enabled": True, "role": "total_value"},
    {"key": "entry_value", "label": "Valor da entrada", "type": "money", "required": False, "enabled": True, "role": "entry_value"},
    {"key": "installment_count", "label": "Quantidade de parcelas", "type": "number", "required": True, "enabled": True, "role": "installment_count"},
    {"key": "first_due_date", "label": "Primeiro vencimento", "type": "date", "required": True, "enabled": True, "role": "first_due_date"},
    {"key": "operator", "label": "Operador", "type": "text", "required": False, "enabled": True, "role": "operator"},
    {"key": "notes", "label": "Observacoes", "type": "textarea", "required": False, "enabled": True, "role": "notes"},
]

PROFILE_FIELD_OVERRIDES = {
    "alpha": {
        "identifier": "DEBIT ID",
        "client": "Cliente",
        "document": "CPF/CNPJ",
    },
    "beta": {
        "identifier": "SUITID",
        "client": "Cliente",
        "process": "Processo",
    },
}

REQUIRED_HEADERS = [
    "DEBIT ID",
    "CPF/CNPJ",
    "CLIENTE",
    "VALOR DO ACORDO",
    "H.O",
    "PARCELAS",
    "DATA DO VENCIMENTO",
    "STATUS",
    "OBS",
    "OPERADOR",
    "TIPO DE ACORDO",
    "ACORDO",
]

OPEN_STATUSES = {"A VENCER", "VENCIDO"}
VALID_STATUSES = OPEN_STATUSES | {"PAGO", "QUEBRA"}


class ColchaoService:
    def __init__(self, data_dir: Path, repo: Any | None = None) -> None:
        self.data_dir = data_dir
        self.repo = repo
        self.config_path = data_dir / "colchao_config.json"
        self.history_path = data_dir / "colchao_history.json"
        self.backup_dir = data_dir / "backups"
        self.backup_dir.mkdir(exist_ok=True)
        self._lock = threading.Lock()
        self._records_cache: dict[tuple[str, str, float], list[dict[str, Any]]] = {}
        self._dashboard_cache: dict[tuple[str, str, float, str], dict[str, Any]] = {}
        self._overdue_checked: dict[str, str] = {}
        if not self.config_path.exists():
            self.config_path.write_text(json.dumps(DEFAULT_CONFIG, ensure_ascii=False, indent=2), encoding="utf-8")

    def get_config(self) -> dict[str, Any]:
        return DEFAULT_CONFIG | self._read_json(self.config_path, {})

    def get_profile_config(self, profile: str = "alpha") -> dict[str, Any]:
        profile = self._profile_id(profile)
        raw = self.get_config()
        base = self._profile_defaults(profile).copy()
        base["name"] = base.get("name") or profile.title()
        if profile == "alpha":
            base["excel_path"] = raw.get("excel_path", base.get("excel_path", ""))
            base["main_sheet"] = raw.get("main_sheet", base.get("main_sheet", ""))
        profile_config = (raw.get("profiles") or {}).get(profile, {})
        for key in ("excel_path", "main_sheet"):
            if key in profile_config:
                base[key] = profile_config[key]
        stored = self.repo.get_colchao_config(profile) if self.repo else None
        if stored:
            base.update(stored)
        base["fields"] = self._normalize_fields(base.get("fields") or [], profile)
        base.setdefault("statuses", ["A VENCER", "VENCIDO", "PAGO", "QUEBRA"])
        return base

    def save_config(self, payload: dict[str, Any], profile: str = "alpha", usuario: str = "sistema") -> dict[str, Any]:
        profile = self._profile_id(payload.get("profile") or profile)
        current = self.get_profile_config(profile)
        database_config = {
            "name": str(payload.get("name", current.get("name") or profile.title())).strip() or profile.title(),
            "fields": self._normalize_fields(payload.get("fields", current.get("fields") or []), profile),
            "statuses": self._normalize_statuses(payload.get("statuses", current.get("statuses") or [])),
        }
        for key in ("excel_path", "main_sheet", "sheet_options"):
            if key in payload:
                database_config[key] = payload[key]
            elif key in current:
                database_config[key] = current[key]
        if self.repo:
            saved = self.repo.save_colchao_config(profile, database_config, usuario)
        else:
            saved = database_config
        config = self.get_config() if self.config_path.exists() else DEFAULT_CONFIG.copy()
        profiles = dict(config.get("profiles") or {})
        profile_config = dict(profiles.get(profile) or {})
        for key in DEFAULT_CONFIG:
            if key in payload:
                profile_config[key] = str(payload[key]).strip()
        profiles[profile] = profile_config
        config["profiles"] = profiles
        if profile == "alpha":
            config["excel_path"] = profile_config.get("excel_path", config.get("excel_path", ""))
            config["main_sheet"] = profile_config.get("main_sheet", config.get("main_sheet", ""))
        self.config_path.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
        self._clear_cache()
        return {**self.get_profile_config(profile), **saved}

    def config_versions(self, profile: str = "alpha") -> list[dict[str, Any]]:
        return self.repo.list_colchao_config_versions(self._profile_id(profile)) if self.repo else []

    def _default_fields(self, profile: str) -> list[dict[str, Any]]:
        fields = copy.deepcopy(DEFAULT_COLCHAO_FIELDS)
        overrides = PROFILE_FIELD_OVERRIDES.get(profile, {})
        for index, field in enumerate(fields):
            field["order"] = index + 1
            field["label"] = overrides.get(field["role"], field["label"])
            if field["role"] == "operator" and profile in {"alpha", "beta"}:
                field["required"] = True
            if profile == "beta" and field["role"] == "document":
                field["enabled"] = False
            if profile == "beta" and field["role"] == "process":
                field["enabled"] = True
                field["required"] = True
        return fields

    def _normalize_fields(self, fields: Any, profile: str) -> list[dict[str, Any]]:
        if not isinstance(fields, list) or not fields:
            return self._default_fields(profile)
        allowed_types = {"text", "number", "money", "date", "select", "textarea"}
        allowed_roles = {field["role"] for field in DEFAULT_COLCHAO_FIELDS}
        normalized = []
        seen_keys = set()
        seen_roles = set()
        for index, raw_field in enumerate(fields):
            if not isinstance(raw_field, dict):
                continue
            key = self._slug(raw_field.get("key") or raw_field.get("label") or "")
            role = str(raw_field.get("role") or key).strip().lower()
            if not key or key in seen_keys or role not in allowed_roles or role in seen_roles:
                continue
            field_type = str(raw_field.get("type") or "text").strip().lower()
            item = {
                "key": key,
                "label": str(raw_field.get("label") or key.replace("_", " ").title()).strip(),
                "type": field_type if field_type in allowed_types else "text",
                "required": bool(raw_field.get("required")),
                "enabled": raw_field.get("enabled") is not False,
                "role": role,
                "order": index + 1,
            }
            options = raw_field.get("options")
            if item["type"] == "select":
                if isinstance(options, str):
                    options = [part.strip() for part in options.split(",")]
                item["options"] = [str(value).strip() for value in (options or []) if str(value).strip()]
            normalized.append(item)
            seen_keys.add(key)
            seen_roles.add(role)
        required_roles = {"identifier", "client", "total_value", "installment_count", "first_due_date"}
        if profile == "beta":
            required_roles.add("process")
        missing = required_roles - seen_roles
        if missing:
            defaults = {field["role"]: field for field in self._default_fields(profile)}
            normalized.extend(defaults[role] for role in required_roles if role in missing)
        required_by_profile = {
            "alpha": {"identifier", "client", "document", "agreement_type", "total_value", "installment_count", "first_due_date", "operator"},
            "beta": {"identifier", "client", "process", "agreement_type", "total_value", "installment_count", "first_due_date", "operator"},
        }.get(profile, required_roles)
        for field in normalized:
            if field["role"] in required_by_profile:
                field["enabled"] = True
                field["required"] = True
            if profile == "beta" and field["role"] == "document":
                field["enabled"] = False
        return [{**field, "order": index + 1} for index, field in enumerate(normalized)]

    @staticmethod
    def _normalize_statuses(statuses: Any) -> list[str]:
        if isinstance(statuses, str):
            statuses = statuses.split(",")
        result = []
        for status in statuses if isinstance(statuses, list) else []:
            value = str(status).strip().upper()
            if value and value not in result:
                result.append(value)
        return result or ["A VENCER", "VENCIDO", "PAGO", "QUEBRA"]

    def records(self, profile: str = "alpha", sheet_name: str = "") -> list[dict[str, Any]]:
        profile = self._profile_id(profile)
        if self.repo:
            rows = self.repo.list_colchao_records(profile, sheet_name)
            result = []
            for row in rows:
                row["__profile"] = profile
                row["__bucket"] = self._bucket(row, profile)
                result.append(row)
            return result
        if not sheet_name:
            rows = []
            for configured_sheet in self._configured_sheets(profile):
                rows.extend(self._excel_records(profile, configured_sheet))
            return rows
        return self._excel_records(profile, sheet_name)

    def sync_from_excel(self, profile: str = "alpha") -> dict[str, Any]:
        profile = self._profile_id(profile)
        if not self.repo:
            return {"ok": True, "profile": profile, "synced": 0, "database": False}
        sheets = self._configured_sheets(profile, prefer_excel=True)
        rows = []
        for sheet_name in sheets:
            rows.extend(self._excel_records(profile, sheet_name))
        payload = [self._db_payload(row, profile) for row in rows]
        synced = self.repo.replace_colchao_records(profile, payload, "excel_sync")
        self._clear_cache()
        return {"ok": True, "profile": profile, "synced": synced, "database": True}

    def _excel_records(self, profile: str = "alpha", sheet_name: str = "") -> list[dict[str, Any]]:
        profile = self._profile_id(profile)
        cache_key = self._cache_key(profile, sheet_name)
        if cache_key in self._records_cache:
            return copy.deepcopy(self._records_cache[cache_key])
        workbook, sheet, _table, headers, header_row, min_col, max_row = self._open_table(data_only=True, profile=profile, sheet_name=sheet_name)
        self._validate_headers(headers, profile)
        result = []
        for row_index in range(header_row + 1, max_row + 1):
            values = [self._cell_value(sheet.cell(row_index, min_col + index).value) for index in range(len(headers))]
            if not any(value not in ("", None) for value in values):
                continue
            record = {headers[index]: values[index] for index in range(len(headers))}
            record["__row_number"] = row_index
            record["__sheet_name"] = sheet.title
            record["__profile"] = profile
            record["__bucket"] = self._bucket(record, profile)
            result.append(record)
        workbook.close()
        self._records_cache[cache_key] = copy.deepcopy(result)
        return result

    def query_records(
        self,
        page: int = 1,
        page_size: int = 100,
        search: str = "",
        operador: str = "",
        status: str = "",
        vencimento: str = "",
        profile: str = "alpha",
        sheet_name: str = "",
        all_records: bool = False,
    ) -> dict[str, Any]:
        profile = self._profile_id(profile)
        self.auto_mark_overdue(profile)
        page = max(1, int(page or 1))
        page_size = min(500, max(25, int(page_size or 100)))
        filters = {
            "search": self._filter_text(search),
            "operador": self._filter_text(operador),
            "status": self._filter_text(status),
            "vencimento": self._filter_text(vencimento),
        }
        rows = []
        for record in self.records(profile, sheet_name):
            searchable = self._filter_text(" ".join(str(value or "") for value in record.values()))
            if filters["search"] and filters["search"] not in searchable:
                continue
            if filters["operador"] and filters["operador"] not in self._filter_text(self._operator_value(record, profile)):
                continue
            if filters["status"] and filters["status"] not in self._filter_text(self._value(record, "STATUS")):
                continue
            if filters["vencimento"] and filters["vencimento"] not in self._filter_text(self._due_value(record, profile)):
                continue
            rows.append(record)
        total = len(rows)
        if all_records:
            return {
                "rows": rows,
                "total": total,
                "page": 1,
                "page_size": total,
                "total_pages": 1,
            }
        total_pages = max(1, (total + page_size - 1) // page_size)
        page = min(page, total_pages)
        start = (page - 1) * page_size
        return {
            "rows": rows[start:start + page_size],
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": total_pages,
        }

    def pendencias(self, profile: str = "alpha") -> list[dict[str, Any]]:
        profile = self._profile_id(profile)
        self.auto_mark_overdue(profile)
        today = date.today()
        rows = []
        source_rows = self.records(profile, "") if self.repo else []
        if not self.repo:
            for sheet_name in self._configured_sheets(profile, prefer_excel=True):
                source_rows.extend(self.records(profile, sheet_name))
        for record in source_rows:
            due = self._parse_date(self._due_value(record, profile))
            status = self._status(record)
            if not due or status not in OPEN_STATUSES:
                continue
            if (status == "VENCIDO" and due <= today) or (status == "A VENCER" and due >= today):
                rows.append(record)
        return sorted(rows, key=lambda row: (self._parse_date(self._due_value(row, profile)) or today, self._client_value(row, profile)))

    def dashboard(self, profile: str = "alpha") -> dict[str, Any]:
        profile = self._profile_id(profile)
        self.auto_mark_overdue(profile)
        today = date.today()
        cache_key = self._cache_key(profile, "") + (today.isoformat(),)
        if cache_key in self._dashboard_cache:
            return copy.deepcopy(self._dashboard_cache[cache_key])
        records = self.records(profile, "") if self.repo else []
        if not self.repo:
            for sheet_name in self._configured_sheets(profile, prefer_excel=True):
                records.extend(self.records(profile, sheet_name))
        open_rows = [row for row in records if self._status(row) in OPEN_STATUSES]
        paid_rows = [row for row in records if self._status(row) == "PAGO"]
        break_rows = [row for row in records if self._status(row) == "QUEBRA"]
        dashboard = {
            "total_registros": len(records),
            "pendencias": len(open_rows),
            "a_vencer_hoje": len([row for row in records if self._bucket(row, profile) == "a_vencer_hoje"]),
            "vencidas": len([row for row in records if self._status(row) == "VENCIDO" or (self._status(row) == "A VENCER" and (self._parse_date(self._due_value(row, profile)) or today) < today)]),
            "quebras": len({(self._id_value(row, profile), self._value(row, "ACORDO")) for row in break_rows}),
            "pagos": len({(self._id_value(row, profile), self._value(row, "ACORDO")) for row in paid_rows}),
            "valor_aberto": sum(self._money_value(row, profile) for row in open_rows),
            "valor_pago": sum(self._money_value(row, profile) for row in paid_rows),
            "clientes_ativos": len({self._id_value(row, profile) for row in open_rows if self._id_value(row, profile)}),
            "ranking_operador": self._ranking(open_rows, profile),
            "updated_at": datetime.now().isoformat(timespec="seconds"),
        }
        self._dashboard_cache[cache_key] = copy.deepcopy(dashboard)
        return dashboard

    def relatorio_csv(self, profile: str = "alpha") -> tuple[str, bytes]:
        profile = self._profile_id(profile)
        self.auto_mark_overdue(profile)
        rows = self.records(profile, "")
        headers = self._report_headers(rows, profile)
        output = io.StringIO(newline="")
        writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_MINIMAL)
        writer.writerow(headers)
        for row in rows:
            writer.writerow([self._csv_value(self._report_value(row, header)) for header in headers])
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"relatorio_colchao_{profile}_{stamp}.csv"
        return filename, ("\ufeff" + output.getvalue()).encode("utf-8")

    def validate(self, profile: str = "alpha") -> dict[str, Any]:
        profile = self._profile_id(profile)
        records = self.records(profile, "") if self.repo else []
        if not self.repo:
            for sheet_name in self._configured_sheets(profile, prefer_excel=True):
                records.extend(self.records(profile, sheet_name))
        errors = []
        for record in records:
            row = record.get("__row_number")
            sheet = record.get("__sheet_name", "")
            if not self._id_value(record, profile):
                errors.append(f"{sheet} linha {row}: identificador vazio.")
            if not self._value(record, "ACORDO"):
                errors.append(f"{sheet} linha {row}: ACORDO vazio.")
            if self._status(record) not in VALID_STATUSES:
                errors.append(f"{sheet} linha {row}: STATUS invalido.")
            if not self._parse_date(self._due_value(record, profile)):
                errors.append(f"{sheet} linha {row}: vencimento invalido.")
        return {"ok": not errors, "errors": errors[:200], "total_errors": len(errors)}

    def _report_headers(self, rows: list[dict[str, Any]], profile: str) -> list[str]:
        base = ["SHEET", "LINHA"]
        preferred = self._profile_defaults(profile)["id_labels"] + [
            "CPF/CNPJ",
            "NOME",
            "CLIENTE",
            "ACORDO",
            "PARCELAS",
            "COND PARCELADAS",
            "VALOR DO ACORDO",
            "CASH",
            "DATA DO VENCIMENTO",
            "MES",
            "STATUS",
            "OBS",
            "OPERADOR",
        ]
        seen = set(base)
        headers = list(base)
        for header in preferred:
            key = self._matching_key(rows, header)
            if key and key not in seen:
                seen.add(key)
                headers.append(key)
        for row in rows:
            for key in row:
                if str(key).startswith("__") or key in seen:
                    continue
                seen.add(key)
                headers.append(key)
        return headers

    def _matching_key(self, rows: list[dict[str, Any]], header: str) -> str:
        wanted = self._normalize_header(header)
        for row in rows:
            for key in row:
                if self._normalize_header(key) == wanted:
                    return key
        return ""

    def _report_value(self, row: dict[str, Any], header: str) -> Any:
        if header == "SHEET":
            return row.get("__sheet_name", "")
        if header == "LINHA":
            return row.get("__row_number", "")
        return row.get(header, "")

    def _csv_value(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, (date, datetime)):
            return value.strftime("%d/%m/%Y")
        return str(value).replace("\r", " ").replace("\n", " ").strip()

    def _normalize_header(self, value: Any) -> str:
        text = unicodedata.normalize("NFD", str(value or ""))
        text = "".join(char for char in text if unicodedata.category(char) != "Mn")
        return "".join(char for char in text.upper() if char.isalnum())

    def update_status(self, row_number: int, status: str, observacao: str, user: str, profile: str = "alpha", sheet_name: str = "") -> dict[str, Any]:
        profile = self._profile_id(profile)
        status = self._normalize_status(status)
        row_number = int(row_number or 0)
        with self._lock:
            backup = self.create_backup(profile)
            records = self.records(profile, sheet_name)
            changed = self._planned_status_changes(records, row_number, status, user, observacao, profile)
            self._write_status_changes_to_database(changed, profile, user)
            self._write_status_changes_with_excel(changed, profile, sheet_name)
            self._log_many(changed)
            self._clear_cache()
        return {"ok": True, "backup": str(backup) if backup else "", "changed": changed}

    def update_status_batch(self, changes: list[dict[str, Any]], user: str, profile: str = "alpha", sheet_name: str = "") -> dict[str, Any]:
        profile = self._profile_id(profile)
        normalized_changes = self._normalize_batch_changes(changes)
        if not normalized_changes:
            return {"ok": True, "backup": None, "changed": []}
        with self._lock:
            backup = self.create_backup(profile)
            records = self.records(profile, sheet_name)
            by_row: dict[int, dict[str, Any]] = {}
            due_by_row: dict[int, dict[str, Any]] = {}
            for item in normalized_changes:
                if item.get("status"):
                    for change in self._planned_status_changes(records, item["row"], item["status"], user, item.get("observacao", ""), profile):
                        by_row[int(change["row"])] = change
                if item.get("vencimento"):
                    due_change = self._planned_due_date_change(records, item["row"], item["vencimento"], user, profile)
                    if due_change:
                        due_by_row[int(due_change["row"])] = due_change
            status_changes = list(by_row.values())
            due_changes = list(due_by_row.values())
            changed = status_changes + due_changes
            self._write_status_changes_to_database(status_changes, profile, user)
            self._write_due_date_changes_to_database(due_changes, profile, user)
            self._write_status_changes_with_excel(status_changes, profile, sheet_name)
            self._write_due_date_changes_with_excel(due_changes, profile, sheet_name)
            self._log_many(changed)
            self._clear_cache()
        return {"ok": True, "backup": str(backup) if backup else "", "changed": changed}

    def clients(self, profile: str = "alpha", search: str = "") -> dict[str, Any]:
        profile = self._profile_id(profile)
        rows = self._profile_records(profile)
        query = self._filter_text(search)
        clients: dict[str, dict[str, Any]] = {}
        for row in rows:
            identifier = str(self._id_value(row, profile) or "").strip()
            client_name = str(self._client_value(row, profile) or "Cliente nao identificado").strip()
            document = str(self._first_value(row, ["CPF/CNPJ", "CPF", "CNPJ"]) or "").strip()
            operator = str(self._operator_value(row, profile) or "Sem operador").strip()
            searchable = self._filter_text(" ".join((identifier, client_name, document, operator)))
            if query and query not in searchable:
                continue
            client_key = self._normalize(identifier) or self._normalize(document) or self._normalize(client_name)
            client = clients.setdefault(client_key, {
                "key": client_key,
                "identifier": identifier,
                "client": client_name,
                "document": document,
                "operator": operator,
                "agreements": {},
            })
            agreement_number = str(self._value(row, "ACORDO") or "1").strip() or "1"
            sheet = str(row.get("__sheet_name") or "")
            agreement_key = f"{sheet}::{agreement_number}"
            agreement = client["agreements"].setdefault(agreement_key, {
                "key": agreement_key,
                "number": agreement_number,
                "sheet": sheet,
                "type": str(self._first_value(row, ["TIPO DE ACORDO", "TIPO", "CONDICAO"]) or "").strip(),
                "operator": operator,
                "installments": [],
            })
            due = self._parse_date(self._due_value(row, profile))
            agreement["installments"].append({
                "row": int(row.get("__row_number") or 0),
                "sheet": sheet,
                "label": str(self._installment_value(row, profile) or "").strip(),
                "due_date": due.isoformat() if due else "",
                "due_date_label": due.strftime("%d/%m/%Y") if due else "Sem data",
                "value": self._money_value(row, profile),
                "status": self._status(row),
                "observation": str(self._first_value(row, ["OBS", "OBSERVACOES", "OBSERVAÇÕES"]) or "").strip(),
            })

        result = []
        today = date.today()
        for client in clients.values():
            agreements = []
            for agreement in client.pop("agreements").values():
                installments = sorted(agreement["installments"], key=lambda item: (item["due_date"] or "9999", item["row"]))
                open_items = [item for item in installments if item["status"] in OPEN_STATUSES]
                paid_items = [item for item in installments if item["status"] == "PAGO"]
                broken_items = [item for item in installments if item["status"] == "QUEBRA"]
                overdue = [item for item in open_items if item["due_date"] and date.fromisoformat(item["due_date"]) < today]
                if open_items:
                    agreement_status = "VENCIDO" if overdue else "ATIVO"
                elif paid_items and len(paid_items) == len(installments):
                    agreement_status = "QUITADO"
                elif broken_items:
                    agreement_status = "QUEBRADO"
                else:
                    agreement_status = "ENCERRADO"
                agreement.update({
                    "status": agreement_status,
                    "installments": installments,
                    "installment_count": len(installments),
                    "open_count": len(open_items),
                    "paid_count": len(paid_items),
                    "broken_count": len(broken_items),
                    "overdue_count": len(overdue),
                    "total_value": sum(item["value"] for item in installments),
                    "open_value": sum(item["value"] for item in open_items),
                    "paid_value": sum(item["value"] for item in paid_items),
                    "next_due_date": min((item["due_date"] for item in open_items if item["due_date"]), default=""),
                })
                agreements.append(agreement)
            agreements.sort(key=lambda item: (item["number"], item["sheet"]), reverse=True)
            all_installments = [item for agreement in agreements for item in agreement["installments"]]
            client.update({
                "agreements": agreements,
                "agreement_count": len(agreements),
                "active_count": sum(agreement["status"] in {"ATIVO", "VENCIDO"} for agreement in agreements),
                "broken_count": sum(agreement["status"] == "QUEBRADO" for agreement in agreements),
                "paid_count": sum(agreement["status"] == "QUITADO" for agreement in agreements),
                "overdue_count": sum(agreement["overdue_count"] for agreement in agreements),
                "open_value": sum(agreement["open_value"] for agreement in agreements),
                "paid_value": sum(agreement["paid_value"] for agreement in agreements),
                "next_due_date": min((agreement["next_due_date"] for agreement in agreements if agreement["next_due_date"]), default=""),
                "installment_count": len(all_installments),
            })
            result.append(client)
        result.sort(key=lambda item: (-item["overdue_count"], item["client"].upper()))
        return {
            "items": result,
            "total": len(result),
            "agreements": sum(item["agreement_count"] for item in result),
            "open_value": sum(item["open_value"] for item in result),
        }

    def preview_due_date_reschedule(self, payload: dict[str, Any], user: str) -> dict[str, Any]:
        profile = self._profile_id(payload.get("profile"))
        return self._plan_due_date_reschedule(self._profile_records(profile), payload, user, profile)

    def reschedule_due_dates(self, payload: dict[str, Any], user: str) -> dict[str, Any]:
        profile = self._profile_id(payload.get("profile"))
        with self._lock:
            plan = self._plan_due_date_reschedule(self._profile_records(profile), payload, user, profile)
            changes = plan["changes"]
            if not changes:
                raise ColchaoError("Nenhuma parcela aberta foi encontrada para reprogramar.")
            backup = self.create_backup(profile)
            self._write_due_date_changes_to_database(changes, profile, user)
            self._write_due_date_changes_with_excel(changes, profile)
            self._log_many(changes)
            self._clear_cache()
        return {**plan, "ok": True, "backup": str(backup) if backup else ""}

    def _plan_due_date_reschedule(self, rows: list[dict[str, Any]], payload: dict[str, Any], user: str, profile: str) -> dict[str, Any]:
        row_number = int(payload.get("row") or 0)
        sheet = str(payload.get("sheet") or "").strip()
        scope = str(payload.get("scope") or "selected").strip().lower()
        mode = str(payload.get("mode") or "schedule").strip().lower()
        new_date = self._parse_date(payload.get("new_date"))
        if row_number <= 0 or not new_date:
            raise ColchaoError("Parcela e nova data sao obrigatorias.")
        if scope not in {"selected", "from_current_month", "from_next_month", "all_open"}:
            raise ColchaoError("Escopo de reprogramacao invalido.")
        if mode not in {"schedule", "day"}:
            raise ColchaoError("Modo de reprogramacao invalido.")
        base = next((row for row in rows if int(row.get("__row_number") or 0) == row_number and (not sheet or str(row.get("__sheet_name") or "") == sheet)), None)
        if not base:
            raise ColchaoError("Parcela selecionada nao foi encontrada.")
        agreement_rows = [
            row for row in rows
            if str(row.get("__sheet_name") or "") == str(base.get("__sheet_name") or "")
            and str(self._id_value(row, profile) or "") == str(self._id_value(base, profile) or "")
            and str(self._value(row, "ACORDO") or "") == str(self._value(base, "ACORDO") or "")
            and self._status(row) in OPEN_STATUSES
        ]
        base_due = self._parse_date(self._due_value(base, profile)) or new_date
        month_start = date(base_due.year, base_due.month, 1)
        next_month_start = self._add_months(month_start, 1)
        if scope == "selected":
            selected = [row for row in agreement_rows if int(row.get("__row_number") or 0) == row_number]
        elif scope == "from_current_month":
            selected = [row for row in agreement_rows if (self._parse_date(self._due_value(row, profile)) or date.max) >= month_start]
        elif scope == "from_next_month":
            selected = [row for row in agreement_rows if (self._parse_date(self._due_value(row, profile)) or date.max) >= next_month_start]
        else:
            selected = agreement_rows
        selected.sort(key=lambda row: (self._parse_date(self._due_value(row, profile)) or date.max, int(row.get("__row_number") or 0)))
        reason = str(payload.get("reason") or "").strip()
        changes = []
        for index, row in enumerate(selected):
            current_due = self._parse_date(self._due_value(row, profile))
            if mode == "schedule":
                target = self._add_months(new_date, index)
            else:
                source = current_due or new_date
                target = date(source.year, source.month, min(new_date.day, calendar.monthrange(source.year, source.month)[1]))
            change = self._planned_due_date_change([row], int(row.get("__row_number") or 0), target, user, profile)
            if not change:
                continue
            cloned = dict(row)
            cloned[change["header"]] = change["depois"]
            change["bucket"] = self._bucket(cloned, profile)
            change["observacao"] = "Reprogramacao de vencimentos pelo sistema gerencial." + (f" Motivo: {reason}" if reason else "")
            changes.append(change)
        return {
            "profile": profile,
            "client": str(self._client_value(base, profile) or "Cliente nao identificado"),
            "identifier": str(self._id_value(base, profile) or ""),
            "agreement": str(self._value(base, "ACORDO") or ""),
            "scope": scope,
            "mode": mode,
            "changes": changes,
            "total": len(changes),
        }

    def auto_break_old_agreements(self, user: str) -> dict[str, Any]:
        with self._lock:
            backup = self.create_backup()
            records = self.records()
            latest_by_debit: dict[str, int] = {}
            for record in records:
                debit_id = str(self._value(record, "DEBIT ID")).strip()
                agreement = self._agreement_number(self._value(record, "ACORDO"))
                if debit_id and self._status(record) in OPEN_STATUSES:
                    latest_by_debit[debit_id] = max(latest_by_debit.get(debit_id, agreement), agreement)
            changed = []
            for record in records:
                debit_id = str(self._value(record, "DEBIT ID")).strip()
                agreement = self._agreement_number(self._value(record, "ACORDO"))
                if debit_id and agreement < latest_by_debit.get(debit_id, agreement) and self._status(record) in OPEN_STATUSES:
                    changed.append(self._status_change_payload(record, int(record["__row_number"]), "QUEBRA", user, "Quebra automatica por acordo posterior."))
            self._write_status_changes_to_database(changed, "alpha", user)
            self._write_status_changes_with_excel(changed)
            self._log_many(changed)
            self._clear_cache()
        return {"ok": True, "backup": str(backup) if backup else "", "changed": changed}

    def auto_mark_overdue(self, profile: str = "alpha", user: str = "sistema", force: bool = False) -> dict[str, Any]:
        profile = self._profile_id(profile)
        today = date.today()
        today_key = today.isoformat()
        if not force and self._overdue_checked.get(profile) == today_key:
            return {"ok": True, "backup": None, "changed": [], "cached": True}
        with self._lock:
            if not force and self._overdue_checked.get(profile) == today_key:
                return {"ok": True, "backup": None, "changed": [], "cached": True}
            sheets = self._configured_sheets(profile, prefer_excel=not bool(self.repo))
            changes_by_sheet: dict[str, list[dict[str, Any]]] = {}
            source_rows = self.records(profile, "") if self.repo else []
            if not self.repo:
                for sheet_name in sheets:
                    source_rows.extend(self.records(profile, sheet_name))
            for record in source_rows:
                sheet_name = str(record.get("__sheet_name") or "")
                due = self._parse_date(self._due_value(record, profile))
                if not due or due >= today or self._status(record) != "A VENCER":
                    continue
                change = self._status_change_payload(
                    record,
                    int(record["__row_number"]),
                    "VENCIDO",
                    user,
                    "Status atualizado automaticamente por vencimento.",
                    profile,
                )
                if change:
                    changes_by_sheet.setdefault(str(record.get("__sheet_name") or sheet_name), []).append(change)
            changed = [change for changes in changes_by_sheet.values() for change in changes]
            if not changed:
                self._overdue_checked[profile] = today_key
                return {"ok": True, "backup": None, "changed": []}
            backup = self.create_backup(profile)
            for sheet_name, changes in changes_by_sheet.items():
                self._write_status_changes_to_database(changes, profile, user)
                self._write_status_changes_with_excel(changes, profile, sheet_name)
            self._log_many(changed)
            self._clear_cache()
            self._overdue_checked[profile] = today_key
        return {"ok": True, "backup": str(backup) if backup else "", "changed": changed}

    def create_agreement(self, payload: dict[str, Any], user: str) -> dict[str, Any]:
        profile = self._profile_id(payload.get("profile"))
        config = self.get_profile_config(profile)
        payload = self._create_payload_from_roles(payload, config)
        self._validate_create(payload, config)
        sheet_name = str(payload.get("sheet") or payload.get("main_sheet") or config.get("main_sheet") or "COLCHAO").strip() or "COLCHAO"
        with self._lock:
            backup = self.create_backup(profile)
            parcelas = int(payload.get("parcelas") or 1)
            total = self._money(payload.get("valor_acordo"))
            parcel_values = self._parcel_values(total, parcelas, payload.get("entrada"))
            start_date = self._parse_date(payload.get("data_vencimento")) or date.today()
            identifier = payload.get("suitid") if profile == "beta" else payload.get("debit_id")
            agreement_number = self._next_agreement_number_from_records(self._profile_records(profile), identifier, profile)
            rows_payload = []
            for index in range(parcelas):
                vencimento = self._add_months(start_date, index)
                if profile == "beta":
                    rows_payload.append(self._beta_agreement_payload(payload, parcel_values[index], index + 1, parcelas, vencimento, agreement_number))
                elif profile == "alpha":
                    rows_payload.append(self._alpha_agreement_payload(payload, parcel_values[index], index + 1, parcelas, vencimento, agreement_number))
                else:
                    rows_payload.append(self._generic_agreement_payload(payload, parcel_values[index], index + 1, parcelas, vencimento, agreement_number))
            if not self.repo:
                info = self._table_info(profile, sheet_name)
                created_rows = self._append_agreement_with_excel(info, rows_payload, profile)
            else:
                db_rows = []
                for raw in rows_payload:
                    record = dict(raw)
                    record["__sheet_name"] = sheet_name
                    record["__row_number"] = 1
                    record["__profile"] = profile
                    db_rows.append(self._db_payload(record, profile))
                created_rows = self.repo.append_colchao_records(profile, db_rows, user)
                if self._has_excel_source(profile):
                    try:
                        info = self._table_info(profile, sheet_name)
                        self._append_agreement_with_excel(info, rows_payload, profile)
                    except (ColchaoError, OSError):
                        pass
            self._clear_cache()
        if profile == "alpha":
            self.auto_break_old_agreements(user)
        return {"ok": True, "rows": created_rows, "agreement": agreement_number, "backup": str(backup) if backup else ""}

    def history(self) -> list[dict[str, Any]]:
        return self._read_json(self.history_path, [])

    def create_backup(self, profile: str = "alpha") -> Path | None:
        if not self._has_excel_source(profile):
            return None
        path = self._excel_path(profile)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        target = self.backup_dir / f"{path.stem}_{stamp}{path.suffix}"
        shutil.copy2(path, target)
        return target

    def _table_info(self, profile: str = "alpha", sheet_name: str = "") -> dict[str, Any]:
        workbook, sheet, table, headers, header_row, min_col, max_row = self._open_table(data_only=True, profile=profile, sheet_name=sheet_name)
        info = {
            "sheet_name": sheet.title,
            "table_name": getattr(table, "displayName", None) or getattr(table, "name", None),
            "headers": headers,
            "header_row": header_row,
            "min_col": min_col,
            "max_row": max_row,
        }
        workbook.close()
        return info

    def _run_excel_write(self, action, profile: str = "alpha"):
        try:
            import pythoncom
            import win32com.client
        except ImportError as exc:
            raise ColchaoError("Alteracao no COLCHAO requer Microsoft Excel e pywin32/win32com instalados neste Python.") from exc
        path = self._excel_path(profile)
        pythoncom.CoInitialize()
        excel = None
        workbook = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.EnableEvents = False
            try:
                excel.AskToUpdateLinks = False
            except Exception:
                pass
            workbook = excel.Workbooks.Open(str(path), UpdateLinks=0, ReadOnly=False)
            result = action(excel, workbook)
            workbook.Save()
            return result
        except Exception as exc:
            raise ColchaoError(f"Falha ao salvar planilha pelo Microsoft Excel: {exc}") from exc
        finally:
            if workbook is not None:
                try:
                    workbook.Close(SaveChanges=False)
                except Exception:
                    pass
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()

    def _excel_sheet_and_table(self, workbook, info: dict[str, Any]):
        sheet = workbook.Worksheets(info["sheet_name"])
        table_name = info.get("table_name")
        table = sheet.ListObjects(table_name) if table_name else sheet.ListObjects(1)
        return sheet, table

    def _excel_table_column(self, table, header: str) -> int:
        normalized = self._normalize(header)
        for index in range(1, table.ListColumns.Count + 1):
            if self._normalize(table.ListColumns(index).Name) == normalized:
                return index
        raise ColchaoError(f"Coluna '{header}' nao encontrada na tabela do Excel.")

    def _write_status_changes_with_excel(self, changes: list[dict[str, Any]], profile: str = "alpha", sheet_name: str = "") -> None:
        if not changes or not self._has_excel_source(profile):
            return
        try:
            self._write_status_changes_with_excel_com(changes, profile, sheet_name)
        except ColchaoError as exc:
            try:
                self._write_status_changes_with_openpyxl(changes, profile, sheet_name)
            except ColchaoError:
                # O banco e a fonte oficial. Falha na copia Excel nao desfaz a operacao.
                return

    def _write_status_changes_with_excel_com(self, changes: list[dict[str, Any]], profile: str = "alpha", sheet_name: str = "") -> None:
        info = self._table_info(profile, sheet_name)

        def action(_excel, workbook):
            sheet, table = self._excel_sheet_and_table(workbook, info)
            status_index = self._excel_table_column(table, "STATUS")
            status_col = table.ListColumns(status_index).Range.Column
            for change in changes:
                sheet.Cells(int(change["row"]), status_col).Value = change["depois"]
            return True

        self._run_excel_write(action, profile)

    def _write_status_changes_with_openpyxl(self, changes: list[dict[str, Any]], profile: str = "alpha", sheet_name: str = "") -> None:
        profile = self._profile_id(profile)
        path = self._excel_path(profile)
        try:
            workbook = load_workbook(path, data_only=False, keep_vba=self._keep_vba(path))
        except Exception as exc:
            raise ColchaoError(f"Falha ao abrir planilha COLCHAO para salvar status: {exc}") from exc

        try:
            grouped: dict[str, list[dict[str, Any]]] = {}
            fallback_sheet = str(sheet_name or self.get_profile_config(profile).get("main_sheet") or "").strip()
            for change in changes:
                target_sheet = str(change.get("sheet") or fallback_sheet or "").strip()
                grouped.setdefault(target_sheet, []).append(change)

            for target_sheet, sheet_changes in grouped.items():
                selected_sheet = target_sheet or workbook.sheetnames[0]
                if selected_sheet not in workbook.sheetnames:
                    raise ColchaoError(f"Aba '{selected_sheet}' nao encontrada para salvar status do colchao.")
                sheet = workbook[selected_sheet]
                table = self._find_table(sheet, profile)
                if table is None:
                    raise ColchaoError(f"Nenhuma tabela formatada foi encontrada na aba '{selected_sheet}'.")
                headers, header_row, min_col, max_row = self._headers(sheet, table.ref)
                status_header = self._header_lookup(headers, "STATUS")
                if not status_header:
                    raise ColchaoError("Coluna STATUS nao encontrada na tabela do colchao.")
                status_col = min_col + headers.index(status_header)
                for change in sheet_changes:
                    row_number = int(change.get("row") or 0)
                    if row_number <= header_row or row_number > max_row:
                        raise ColchaoError(f"Linha {row_number} fora da tabela do colchao na aba '{selected_sheet}'.")
                    sheet.cell(row_number, status_col).value = change.get("depois")
            workbook.save(path)
            self._repair_xlsx_macro_metadata(path)
        except ColchaoError:
            raise
        except Exception as exc:
            raise ColchaoError(f"Falha ao salvar status na planilha COLCHAO: {exc}") from exc
        finally:
            workbook.close()

    def _write_due_date_changes_with_excel(self, changes: list[dict[str, Any]], profile: str, sheet_name: str = "") -> None:
        if not changes or not self._has_excel_source(profile):
            return
        try:
            self._write_due_date_changes_with_excel_com(changes, profile, sheet_name)
        except ColchaoError:
            try:
                self._write_due_date_changes_with_openpyxl(changes, profile, sheet_name)
            except ColchaoError:
                # O banco e a fonte oficial. Falha na copia Excel nao desfaz a operacao.
                return

    def _write_due_date_changes_with_excel_com(self, changes: list[dict[str, Any]], profile: str, sheet_name: str = "") -> None:
        fallback_sheet = str(sheet_name or self.get_profile_config(profile).get("main_sheet") or "").strip()

        def action(_excel, workbook):
            destinations: dict[str, tuple[Any, Any]] = {}
            for change in changes:
                selected_sheet = str(change.get("sheet") or fallback_sheet).strip()
                if selected_sheet not in destinations:
                    info = self._table_info(profile, selected_sheet)
                    destinations[selected_sheet] = self._excel_sheet_and_table(workbook, info)
                sheet, table = destinations[selected_sheet]
                column_index = self._excel_table_column(table, str(change.get("header") or "DATA DO VENCIMENTO"))
                column = table.ListColumns(column_index).Range.Column
                cell = sheet.Cells(int(change["row"]), column)
                cell.Value = datetime.combine(self._parse_date(change["depois"]), datetime.min.time())
                cell.NumberFormat = "dd/mm/yyyy"
            return True

        self._run_excel_write(action, profile)

    def _write_due_date_changes_with_openpyxl(self, changes: list[dict[str, Any]], profile: str, sheet_name: str = "") -> None:
        path = self._excel_path(profile)
        try:
            workbook = load_workbook(path, data_only=False, keep_vba=self._keep_vba(path))
        except Exception as exc:
            raise ColchaoError(f"Falha ao abrir planilha COLCHAO para salvar vencimento: {exc}") from exc
        try:
            fallback_sheet = str(sheet_name or self.get_profile_config(profile).get("main_sheet") or "").strip()
            for change in changes:
                selected_sheet = str(change.get("sheet") or fallback_sheet or workbook.sheetnames[0]).strip()
                if selected_sheet not in workbook.sheetnames:
                    raise ColchaoError(f"Aba '{selected_sheet}' nao encontrada para salvar vencimento do colchao.")
                sheet = workbook[selected_sheet]
                table = self._find_table(sheet, profile)
                if table is None:
                    raise ColchaoError(f"Nenhuma tabela formatada foi encontrada na aba '{selected_sheet}'.")
                headers, header_row, min_col, max_row = self._headers(sheet, table.ref)
                header = self._header_lookup(headers, str(change.get("header") or "DATA DO VENCIMENTO"))
                if not header:
                    raise ColchaoError("Coluna de vencimento nao encontrada na tabela do colchao.")
                row_number = int(change.get("row") or 0)
                if row_number <= header_row or row_number > max_row:
                    raise ColchaoError(f"Linha {row_number} fora da tabela do colchao na aba '{selected_sheet}'.")
                cell = sheet.cell(row_number, min_col + headers.index(header))
                cell.value = self._parse_date(change["depois"])
                cell.number_format = "dd/mm/yyyy"
            workbook.save(path)
            self._repair_xlsx_macro_metadata(path)
        except ColchaoError:
            raise
        except Exception as exc:
            raise ColchaoError(f"Falha ao salvar vencimento na planilha COLCHAO: {exc}") from exc
        finally:
            workbook.close()

    def _append_agreement_with_excel(self, info: dict[str, Any], rows_payload: list[dict[str, Any]], profile: str = "alpha") -> list[int]:
        if not rows_payload:
            return []
        try:
            return self._append_agreement_with_excel_com(info, rows_payload, profile)
        except ColchaoError as exc:
            try:
                return self._append_agreement_with_openpyxl(info, rows_payload, profile)
            except ColchaoError as fallback_exc:
                raise ColchaoError(f"{fallback_exc} Tentativa via Microsoft Excel tambem falhou: {exc}") from fallback_exc

    def _append_agreement_with_excel_com(self, info: dict[str, Any], rows_payload: list[dict[str, Any]], profile: str = "alpha") -> list[int]:
        rate = "22%" if self._profile_id(profile) == "beta" else "8%"
        value_labels = ["CASH", "VALOR DO ACORDO"] if self._profile_id(profile) == "beta" else ["VALOR DO ACORDO", "CASH"]

        def action(_excel, workbook):
            _sheet, table = self._excel_sheet_and_table(workbook, info)
            column_map = {self._normalize(table.ListColumns(index).Name): index for index in range(1, table.ListColumns.Count + 1)}
            created_rows = []
            for row_payload in rows_payload:
                list_row = table.ListRows.Add()
                created_rows.append(int(list_row.Range.Row))
                for label, value in row_payload.items():
                    column_index = column_map.get(self._normalize(label))
                    if column_index:
                        list_row.Range.Cells(1, column_index).Value = value
                ho_index = column_map.get(self._normalize("H.O"))
                value_index = next((column_map.get(self._normalize(label)) for label in value_labels if column_map.get(self._normalize(label))), None)
                if ho_index and value_index:
                    value_cell = list_row.Range.Cells(1, value_index)
                    value_address = f"{get_column_letter(int(value_cell.Column))}{int(value_cell.Row)}"
                    list_row.Range.Cells(1, ho_index).Formula = f"={value_address}*{rate}"
                if self._profile_id(profile) == "beta":
                    month_index = column_map.get(self._normalize("MÊS")) or column_map.get(self._normalize("MES"))
                    expiration_index = column_map.get(self._normalize("MÊS DE EXPIRAÇÃO")) or column_map.get(self._normalize("MES DE EXPIRACAO"))
                    if month_index and expiration_index:
                        month_cell = list_row.Range.Cells(1, month_index)
                        month_address = f"{get_column_letter(int(month_cell.Column))}{int(month_cell.Row)}"
                        list_row.Range.Cells(1, expiration_index).FormulaLocal = f"=DATA(ANO({month_address});MÊS({month_address});DIA({month_address})+7)"
            return created_rows

        return self._run_excel_write(action, profile)

    def _append_agreement_with_openpyxl(self, info: dict[str, Any], rows_payload: list[dict[str, Any]], profile: str = "alpha") -> list[int]:
        profile = self._profile_id(profile)
        path = self._excel_path(profile)
        try:
            workbook = load_workbook(path, data_only=False, keep_vba=self._keep_vba(path))
        except Exception as exc:
            raise ColchaoError(f"Falha ao abrir planilha COLCHAO para cadastrar acordo: {exc}") from exc

        try:
            sheet_name = str(info.get("sheet_name") or "").strip() or workbook.sheetnames[0]
            if sheet_name not in workbook.sheetnames:
                raise ColchaoError(f"Aba '{sheet_name}' nao encontrada para cadastrar acordo.")
            sheet = workbook[sheet_name]
            table_name = str(info.get("table_name") or "").strip()
            table = sheet.tables.get(table_name) if table_name else None
            if table is None:
                table = self._find_table(sheet, profile)
            if table is None:
                raise ColchaoError(f"Nenhuma tabela formatada foi encontrada na aba '{sheet_name}'.")

            headers, header_row, min_col, max_row = self._headers(sheet, table.ref)
            _min_col, _min_row, max_col, _max_row = range_boundaries(table.ref)
            column_map = {self._normalize(header): index for index, header in enumerate(headers)}
            value_labels = ["CASH", "VALOR DO ACORDO"] if profile == "beta" else ["VALOR DO ACORDO", "CASH"]
            rate = "0.22" if profile == "beta" else "0.08"
            created_rows: list[int] = []
            last_data_row = max_row

            for row_payload in rows_payload:
                new_row = last_data_row + 1
                created_rows.append(new_row)
                self._copy_row_style(sheet, last_data_row, new_row, min_col, max_col)
                for label, value in row_payload.items():
                    column_index = column_map.get(self._normalize(label))
                    if column_index is not None:
                        sheet.cell(new_row, min_col + column_index).value = value

                ho_index = column_map.get(self._normalize("H.O"))
                value_index = next((column_map.get(self._normalize(label)) for label in value_labels if column_map.get(self._normalize(label)) is not None), None)
                if ho_index is not None and value_index is not None:
                    value_address = f"{get_column_letter(min_col + value_index)}{new_row}"
                    sheet.cell(new_row, min_col + ho_index).value = f"={value_address}*{rate}"

                if profile == "beta":
                    month_index = column_map.get(self._normalize("MÊS")) or column_map.get(self._normalize("MES"))
                    expiration_index = column_map.get(self._normalize("MÊS DE EXPIRAÇÃO")) or column_map.get(self._normalize("MES DE EXPIRACAO"))
                    if month_index is not None and expiration_index is not None:
                        month_address = f"{get_column_letter(min_col + month_index)}{new_row}"
                        sheet.cell(new_row, min_col + expiration_index).value = f"=DATE(YEAR({month_address}),MONTH({month_address}),DAY({month_address})+7)"

                last_data_row = new_row

            table.ref = f"{get_column_letter(min_col)}{header_row}:{get_column_letter(max_col)}{last_data_row}"
            if getattr(table, "autoFilter", None) is not None:
                table.autoFilter.ref = table.ref
            workbook.save(path)
            self._repair_xlsx_macro_metadata(path)
            return created_rows
        except ColchaoError:
            raise
        except Exception as exc:
            raise ColchaoError(f"Falha ao cadastrar acordo na planilha COLCHAO: {exc}") from exc
        finally:
            workbook.close()

    def _copy_row_style(self, sheet, source_row: int, target_row: int, min_col: int, max_col: int) -> None:
        if source_row <= 0 or target_row <= 0:
            return
        try:
            sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
        except Exception:
            pass
        for col in range(min_col, max_col + 1):
            source = sheet.cell(source_row, col)
            target = sheet.cell(target_row, col)
            if source.has_style:
                target._style = copy.copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
            if source.alignment:
                target.alignment = copy.copy(source.alignment)
            if source.protection:
                target.protection = copy.copy(source.protection)

    def _planned_status_changes(self, records: list[dict[str, Any]], row_number: int, status: str, user: str, observacao: str, profile: str = "alpha") -> list[dict[str, Any]]:
        base = next((record for record in records if int(record.get("__row_number") or 0) == row_number), None)
        if not base:
            raise ColchaoError(f"Linha {row_number} nao encontrada na tabela.")
        if status != "QUEBRA":
            payload = self._status_change_payload(base, row_number, status, user, observacao, profile)
            return [payload] if payload else []
        identifier = self._id_value(base, profile)
        acordo = str(self._value(base, "ACORDO"))
        changes = []
        for record in records:
            record_row = int(record.get("__row_number") or 0)
            if self._id_value(record, profile) == identifier and str(self._value(record, "ACORDO")) == acordo and self._status(record) in OPEN_STATUSES:
                payload = self._status_change_payload(record, record_row, "QUEBRA", user, observacao, profile)
                if payload:
                    changes.append(payload)
        return changes

    def _status_change_payload(self, record: dict[str, Any], row_number: int, status: str, user: str, observacao: str, profile: str = "alpha") -> dict[str, Any] | None:
        before = self._status(record)
        if before == status:
            return None
        return {
            "data_hora": datetime.now().isoformat(timespec="seconds"),
            "usuario": user,
            "debit_id": self._id_value(record, profile),
            "cliente": self._client_value(record, profile),
            "acordo": self._value(record, "ACORDO"),
            "parcela": self._installment_value(record, profile),
            "row": row_number,
            "sheet": record.get("__sheet_name") or "",
            "antes": before,
            "depois": status,
            "observacao": observacao or "",
        }

    def _planned_due_date_change(self, records: list[dict[str, Any]], row_number: int, value: Any, user: str, profile: str) -> dict[str, Any] | None:
        record = next((item for item in records if int(item.get("__row_number") or 0) == row_number), None)
        if not record:
            raise ColchaoError(f"Linha {row_number} nao encontrada na tabela.")
        due = self._parse_date(value)
        if not due:
            raise ColchaoError("Data de vencimento invalida.")
        before = self._parse_date(self._due_value(record, profile))
        if before == due:
            return None
        header = next(
            (
                key for label in self._profile_defaults(profile)["due_labels"]
                for key in record.keys() if self._normalize(key) == self._normalize(label)
            ),
            "DATA DO VENCIMENTO",
        )
        return {
            "data_hora": datetime.now().isoformat(timespec="seconds"),
            "usuario": user,
            "debit_id": self._id_value(record, profile),
            "cliente": self._client_value(record, profile),
            "acordo": self._value(record, "ACORDO"),
            "parcela": self._installment_value(record, profile),
            "row": row_number,
            "sheet": record.get("__sheet_name") or "",
            "campo": header,
            "header": header,
            "antes": before.strftime("%d/%m/%Y") if before else "",
            "depois": due.strftime("%d/%m/%Y"),
            "observacao": "Vencimento alterado pelo sistema gerencial.",
        }

    def _write_status_changes_to_database(self, changes: list[dict[str, Any]], profile: str, user: str) -> None:
        if not self.repo:
            return
        for change in changes:
            sheet_name = str(change.get("sheet") or self.get_profile_config(profile).get("main_sheet") or "")
            row_number = int(change.get("row") or 0)
            if row_number <= 0 or not sheet_name:
                continue
            record = {
                "STATUS": change.get("depois"),
                "DATA DO VENCIMENTO": "",
            }
            bucket = self._bucket(record, profile)
            try:
                current = next(
                    (
                        row for row in self.repo.list_colchao_records(profile, sheet_name)
                        if int(row.get("__row_number") or 0) == row_number
                    ),
                    None,
                )
                if current:
                    current["STATUS"] = change.get("depois")
                    bucket = self._bucket(current, profile)
            except Exception:
                pass
            self.repo.update_colchao_status(
                profile,
                sheet_name,
                row_number,
                str(change.get("depois") or ""),
                str(change.get("observacao") or ""),
                bucket,
                user,
            )

    def _write_due_date_changes_to_database(self, changes: list[dict[str, Any]], profile: str, user: str) -> None:
        if not self.repo:
            return
        normalized = []
        default_sheet = str(self.get_profile_config(profile).get("main_sheet") or "")
        for change in changes:
            item = dict(change)
            item["sheet"] = str(item.get("sheet") or default_sheet)
            if int(item.get("row") or 0) > 0 and item["sheet"]:
                normalized.append(item)
        if normalized:
            try:
                self.repo.update_colchao_due_dates_batch(profile, normalized, user)
            except ValueError as exc:
                raise ColchaoError(str(exc)) from exc

    def _db_payload(self, record: dict[str, Any], profile: str) -> dict[str, Any]:
        raw = {key: value for key, value in record.items() if not str(key).startswith("__")}
        return {
            "__row_number": int(record.get("__row_number") or 0),
            "__sheet_name": str(record.get("__sheet_name") or ""),
            "__identifier": str(self._id_value(record, profile) or ""),
            "__cliente": str(self._client_value(record, profile) or ""),
            "__cpf_cnpj": str(self._first_value(record, ["CPF/CNPJ", "CPF", "CNPJ"]) or ""),
            "__acordo": str(self._value(record, "ACORDO") or ""),
            "__parcela": str(self._installment_value(record, profile) or ""),
            "__valor": self._money_value(record, profile),
            "__vencimento": str(self._due_value(record, profile) or ""),
            "__status": self._status(record),
            "__observacao": str(self._first_value(record, ["OBS", "OBSERVACOES", "OBSERVAÇÕES"]) or ""),
            "__operador": str(self._operator_value(record, profile) or ""),
            "__bucket": self._bucket(record, profile),
            "__raw": raw,
        }

    def _log_many(self, changes: list[dict[str, Any]]) -> None:
        for change in changes:
            self._log(change)

    def _open_table(self, data_only: bool, profile: str = "alpha", sheet_name: str = ""):
        profile = self._profile_id(profile)
        path = self._excel_path(profile)
        try:
            workbook = load_workbook(path, data_only=data_only, keep_vba=self._keep_vba(path))
        except Exception as exc:
            raise ColchaoError(f"Falha ao abrir planilha COLCHAO: {exc}") from exc
        config = self.get_profile_config(profile)
        selected_sheet = sheet_name or config.get("main_sheet") or workbook.sheetnames[0]
        if selected_sheet not in workbook.sheetnames:
            workbook.close()
            raise ColchaoError(f"Aba '{selected_sheet}' nao encontrada.")
        sheet = workbook[selected_sheet]
        table = self._find_table(sheet, profile)
        if table is None:
            workbook.close()
            raise ColchaoError("Nenhuma tabela formatada foi encontrada na aba do colchao.")
        headers, header_row, min_col, max_row = self._headers(sheet, table.ref)
        return workbook, sheet, table, headers, header_row, min_col, max_row

    def _configured_sheets(self, profile: str = "alpha", prefer_excel: bool = False) -> list[str]:
        profile = self._profile_id(profile)
        config = self.get_profile_config(profile)
        sheets = [str(sheet or "").strip() for sheet in (config.get("sheet_options") or []) if str(sheet or "").strip()]
        if sheets:
            return sheets
        main_sheet = str(config.get("main_sheet") or "").strip()
        if main_sheet:
            return [main_sheet]
        if self.repo and not prefer_excel:
            rows = self.repo.list_colchao_records(profile, "")
            db_sheets = sorted({str(row.get("__sheet_name") or "").strip() for row in rows if str(row.get("__sheet_name") or "").strip()})
            return db_sheets
        return self._discover_table_sheets(profile)

    def _discover_table_sheets(self, profile: str = "alpha") -> list[str]:
        profile = self._profile_id(profile)
        path = self._excel_path(profile)
        try:
            workbook = load_workbook(path, data_only=True, keep_vba=self._keep_vba(path))
        except Exception as exc:
            raise ColchaoError(f"Falha ao abrir planilha COLCHAO: {exc}") from exc
        try:
            sheets = []
            for sheet in workbook.worksheets:
                if self._find_table(sheet, profile) is not None:
                    sheets.append(sheet.title)
            return sheets or [workbook.sheetnames[0]]
        finally:
            workbook.close()

    def excel_path(self, profile: str = "alpha") -> Path:
        return self._excel_path(profile)

    def _excel_path(self, profile: str = "alpha") -> Path:
        raw = str(self.get_profile_config(profile).get("excel_path") or "").strip()
        if not raw:
            raise ColchaoError("Configure o caminho da planilha do COLCHAO.")
        for candidate in self._path_candidates(raw):
            path = Path(candidate)
            path = path if path.is_absolute() else self.data_dir.parent / path
            if path.exists():
                return path
        raise ColchaoError(f"Arquivo Excel nao encontrado: {raw}")

    def _keep_vba(self, path: Path) -> bool:
        return path.suffix.lower() in {".xlsm", ".xltm"}

    def _repair_xlsx_macro_metadata(self, path: Path) -> None:
        if path.suffix.lower() != ".xlsx" or not path.exists() or not zipfile.is_zipfile(path):
            return
        content_types = "[Content_Types].xml"
        workbook_rels = "xl/_rels/workbook.xml.rels"
        macro_content_type = "application/vnd.ms-excel.sheet.macroEnabled.main+xml"
        sheet_content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"
        vba_rel_type = "http://schemas.microsoft.com/office/2006/relationships/vbaProject"

        with zipfile.ZipFile(path, "r") as zin:
            names = zin.namelist()
            has_vba_project = "xl/vbaProject.bin" in names
            if has_vba_project:
                return
            content_xml = zin.read(content_types).decode("utf-8") if content_types in names else ""
            rels_xml = zin.read(workbook_rels).decode("utf-8") if workbook_rels in names else ""
            needs_repair = macro_content_type in content_xml or vba_rel_type in rels_xml or "Target=\"vbaProject.bin\"" in rels_xml
            if not needs_repair:
                return

            temp_path = path.with_suffix(f"{path.suffix}.tmp")
            with zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    if item.filename == "xl/vbaProject.bin":
                        continue
                    data = zin.read(item.filename)
                    if item.filename == content_types:
                        ET.register_namespace("", "http://schemas.openxmlformats.org/package/2006/content-types")
                        root = ET.fromstring(data)
                        ns = "http://schemas.openxmlformats.org/package/2006/content-types"
                        for override in root.findall(f"{{{ns}}}Override"):
                            if override.attrib.get("PartName") == "/xl/workbook.xml" and override.attrib.get("ContentType") == macro_content_type:
                                override.set("ContentType", sheet_content_type)
                        data = ET.tostring(root, encoding="utf-8", xml_declaration=False)
                    elif item.filename == workbook_rels:
                        ET.register_namespace("", "http://schemas.openxmlformats.org/package/2006/relationships")
                        root = ET.fromstring(data)
                        for rel in list(root):
                            if rel.attrib.get("Type") == vba_rel_type or rel.attrib.get("Target") == "vbaProject.bin":
                                root.remove(rel)
                        data = ET.tostring(root, encoding="utf-8", xml_declaration=False)
                    zout.writestr(item, data)
        temp_path.replace(path)

    def _cache_key(self, profile: str = "alpha", sheet_name: str = "") -> tuple[str, str, float]:
        if self.repo:
            return (f"database:{self._profile_id(profile)}", str(sheet_name or ""), 0.0)
        path = self._excel_path(profile)
        config = self.get_profile_config(profile)
        selected_sheet = sheet_name or str(config.get("main_sheet") or "")
        return (str(path.resolve()), selected_sheet, path.stat().st_mtime)

    def _clear_cache(self) -> None:
        self._records_cache.clear()
        self._dashboard_cache.clear()
        self._overdue_checked.clear()

    def _path_candidates(self, raw: str) -> list[str]:
        candidates = [raw]
        current = raw
        for _ in range(3):
            try:
                current = current.encode("latin1").decode("utf-8")
            except UnicodeError:
                break
            if current not in candidates:
                candidates.append(current)
        return candidates

    def _find_table(self, sheet, profile: str = "alpha"):
        tables = getattr(sheet, "tables", {})
        best = None
        best_score = -1
        required = self._required_headers(profile)
        for table in tables.values():
            if not hasattr(table, "ref"):
                table = tables[table]
            headers, *_ = self._headers(sheet, table.ref)
            normalized = {self._normalize(header) for header in headers}
            score = len([header for header in required if self._normalize(header) in normalized])
            if score > best_score:
                best = table
                best_score = score
        return best

    def _headers(self, sheet, table_ref: str):
        min_col, min_row, max_col, max_row = range_boundaries(table_ref)
        headers = []
        for col in range(min_col, max_col + 1):
            value = sheet.cell(min_row, col).value
            headers.append(str(value).strip() if value not in (None, "") else f"COLUNA {col}")
        return headers, min_row, min_col, max_row

    def _validate_headers(self, headers: list[str], profile: str = "alpha") -> None:
        normalized = {self._normalize(header) for header in headers}
        missing = [header for header in self._required_headers(profile) if self._normalize(header) not in normalized]
        if missing:
            raise ColchaoError("Colunas obrigatorias ausentes: " + ", ".join(missing))

    def _row_record(self, sheet, headers: list[str], row_number: int, min_col: int) -> dict[str, Any]:
        return {header: self._cell_value(sheet.cell(row_number, min_col + index).value) for index, header in enumerate(headers)}

    def _set_status(self, sheet, headers: list[str], min_col: int, row_number: int, status: str, user: str, observacao: str) -> dict[str, Any]:
        record = self._row_record(sheet, headers, row_number, min_col)
        status_header = self._header_lookup(headers, "STATUS")
        if not status_header:
            raise ColchaoError("Coluna STATUS nao encontrada.")
        before = self._status(record)
        sheet.cell(row_number, min_col + headers.index(status_header)).value = status
        change = {
            "data_hora": datetime.now().isoformat(timespec="seconds"),
            "usuario": user,
            "debit_id": self._value(record, "DEBIT ID"),
            "cliente": self._value(record, "CLIENTE"),
            "acordo": self._value(record, "ACORDO"),
            "parcela": self._value(record, "PARCELAS"),
            "row": row_number,
            "antes": before,
            "depois": status,
            "observacao": observacao or "",
        }
        self._log(change)
        return change

    def _apply_quebra(self, sheet, headers, min_col, header_row, max_row, base_record, user, observacao):
        debit_id = self._value(base_record, "DEBIT ID")
        acordo = self._value(base_record, "ACORDO")
        changed = []
        for row_number in range(header_row + 1, max_row + 1):
            record = self._row_record(sheet, headers, row_number, min_col)
            if self._value(record, "DEBIT ID") == debit_id and str(self._value(record, "ACORDO")) == str(acordo) and self._status(record) in OPEN_STATUSES:
                changed.append(self._set_status(sheet, headers, min_col, row_number, "QUEBRA", user, observacao))
        return changed

    def _create_payload_from_roles(self, payload: dict[str, Any], config: dict[str, Any]) -> dict[str, Any]:
        values = payload.get("values") if isinstance(payload.get("values"), dict) else {}
        result = dict(payload)
        role_values = {}
        for field in config.get("fields") or []:
            role = str(field.get("role") or "").strip().lower()
            key = str(field.get("key") or "").strip()
            if role and key and key in values:
                role_values[role] = values[key]
        mappings = {
            "identifier": ("debit_id", "suitid"),
            "client": ("cliente",),
            "document": ("cpf_cnpj",),
            "process": ("processo",),
            "total_value": ("valor_acordo",),
            "entry_value": ("entrada",),
            "installment_count": ("parcelas",),
            "first_due_date": ("data_vencimento",),
            "operator": ("operador",),
            "agreement_type": ("tipo_acordo",),
            "notes": ("obs",),
        }
        for role, target_keys in mappings.items():
            if role not in role_values:
                continue
            for target_key in target_keys:
                result[target_key] = role_values[role]
        result.setdefault("entrada", 0)
        result.setdefault("parcelas", 1)
        try:
            installment_count = int(result.get("parcelas") or 1)
        except (TypeError, ValueError):
            installment_count = 0
        result.setdefault("tipo_acordo", "A VISTA" if installment_count == 1 else "PARCELADO")
        return result

    def _validate_create(self, payload: dict[str, Any], config: dict[str, Any] | None = None) -> None:
        profile = self._profile_id(payload.get("profile"))
        if profile not in {"alpha", "beta"}:
            required = []
            role_map = {
                "identifier": "debit_id", "client": "cliente", "document": "cpf_cnpj", "process": "processo",
                "total_value": "valor_acordo", "entry_value": "entrada",
                "installment_count": "parcelas", "first_due_date": "data_vencimento",
                "operator": "operador", "agreement_type": "tipo_acordo", "notes": "obs",
            }
            for field in (config or {}).get("fields") or []:
                if field.get("enabled") is not False and field.get("required") and field.get("role") in role_map:
                    required.append(role_map[field["role"]])
        elif profile == "beta":
            required = ["processo", "suitid", "cliente", "valor_acordo", "parcelas", "data_vencimento", "operador"]
        else:
            required = ["debit_id", "cpf_cnpj", "cliente", "valor_acordo", "parcelas", "data_vencimento", "operador", "tipo_acordo"]
        for key in required:
            if str(payload.get(key, "")).strip() == "":
                raise ColchaoError(f"Campo obrigatorio ausente: {key}.")
        total = self._money(payload.get("valor_acordo"))
        parcelas = int(payload.get("parcelas") or 0)
        if total <= 0:
            raise ColchaoError("Valor total do acordo deve ser maior que zero.")
        if parcelas <= 0:
            raise ColchaoError("Quantidade de parcelas deve ser maior que zero.")

    def _alpha_agreement_payload(self, payload: dict[str, Any], parcel_value: float, parcel_number: int, parcelas: int, vencimento: date, agreement_number: int) -> dict[str, Any]:
        return {
            "DEBIT ID": payload.get("debit_id"),
            "CPF/CNPJ": payload.get("cpf_cnpj"),
            "CLIENTE": payload.get("cliente"),
            "VALOR DO ACORDO": parcel_value,
            "PARCELAS": f"{parcel_number}/{parcelas}",
            "DATA DO VENCIMENTO": vencimento.strftime("%d/%m/%Y"),
            "STATUS": "A VENCER",
            "OBS": payload.get("obs", ""),
            "OPERADOR": payload.get("operador", ""),
            "TIPO DE ACORDO": payload.get("tipo_acordo", ""),
            "ACORDO": agreement_number,
        }

    def _beta_agreement_payload(self, payload: dict[str, Any], parcel_value: float, parcel_number: int, parcelas: int, vencimento: date, agreement_number: int) -> dict[str, Any]:
        return {
            "NÚMERO DO PROCESSO": payload.get("processo"),
            "PROCESSO": payload.get("processo"),
            "SUITID": payload.get("suitid"),
            "SUIT": payload.get("suitid"),
            "NOME": payload.get("cliente"),
            "CASH": parcel_value,
            "COND PARCELAS": f"{parcel_number}/{parcelas}",
            "COND PARCELADAS": f"{parcel_number}/{parcelas}",
            "STATUS": "A VENCER",
            "MÊS": datetime(vencimento.year, vencimento.month, vencimento.day),
            "MES": datetime(vencimento.year, vencimento.month, vencimento.day),
            "OBSERVACOES": payload.get("obs", ""),
            "OBSERVAÇÕES": payload.get("obs", ""),
            "OPERADORES": payload.get("operador", ""),
            "OPERADOR": payload.get("operador", ""),
            "ACORDO": agreement_number,
        }

    def _generic_agreement_payload(self, payload: dict[str, Any], parcel_value: float, parcel_number: int, parcelas: int, vencimento: date, agreement_number: int) -> dict[str, Any]:
        return {
            "IDENTIFICADOR": payload.get("debit_id"),
            "CLIENTE": payload.get("cliente"),
            "CPF/CNPJ": payload.get("cpf_cnpj", ""),
            "VALOR DO ACORDO": parcel_value,
            "PARCELAS": f"{parcel_number}/{parcelas}",
            "DATA DO VENCIMENTO": vencimento.strftime("%d/%m/%Y"),
            "STATUS": "A VENCER",
            "OBS": payload.get("obs", ""),
            "OPERADOR": payload.get("operador", ""),
            "TIPO DE ACORDO": payload.get("tipo_acordo", ""),
            "ACORDO": agreement_number,
        }

    def _parcel_values(self, total: float, parcelas: int, entry_value: Any) -> list[float]:
        entry = round(self._money(entry_value), 2)
        if entry < 0:
            raise ColchaoError("Valor da entrada nao pode ser negativo.")
        total_cents = int(round(total * 100))
        entry_cents = int(round(entry * 100))
        if parcelas == 1:
            return [total_cents / 100]
        if entry_cents > total_cents:
            raise ColchaoError("Valor da entrada excede o valor total do acordo.")
        remaining_count = parcelas - 1
        remaining_cents = total_cents - entry_cents
        monthly_cents = remaining_cents // remaining_count
        values_cents = [entry_cents] + [monthly_cents] * remaining_count
        values_cents[-1] += remaining_cents - (monthly_cents * remaining_count)
        if sum(values_cents) != total_cents:
            raise ColchaoError("A soma das parcelas nao fecha com o valor total do acordo.")
        return [value / 100 for value in values_cents]

    def _next_agreement_number(self, sheet, headers: list[str], header_row: int, min_col: int, max_row: int, debit_id: Any) -> int:
        target_debit = str(debit_id or "").strip()
        highest = 0
        for row_number in range(header_row + 1, max_row + 1):
            record = self._row_record(sheet, headers, row_number, min_col)
            if str(self._value(record, "DEBIT ID")).strip() == target_debit:
                highest = max(highest, self._agreement_number(self._value(record, "ACORDO")))
        return highest + 1

    def _next_agreement_number_from_records(self, records: list[dict[str, Any]], identifier: Any, profile: str = "alpha") -> int:
        target = str(identifier or "").strip()
        highest = 0
        for record in records:
            if str(self._id_value(record, profile)).strip() == target:
                highest = max(highest, self._agreement_number(self._value(record, "ACORDO")))
        return highest + 1

    def _profile_records(self, profile: str) -> list[dict[str, Any]]:
        rows = []
        for sheet_name in self._configured_sheets(profile):
            rows.extend(self.records(profile, sheet_name))
        return rows

    def _normalize_batch_changes(self, changes: list[dict[str, Any]]) -> list[dict[str, Any]]:
        if not isinstance(changes, list):
            raise ColchaoError("Lista de alteracoes invalida.")
        by_row: dict[int, dict[str, Any]] = {}
        for item in changes:
            row_number = int(item.get("row", 0) if isinstance(item, dict) else 0)
            if row_number <= 0:
                continue
            normalized = {"row": row_number, "observacao": str(item.get("observacao", ""))}
            if str(item.get("status") or "").strip():
                normalized["status"] = self._normalize_status(item["status"])
            if str(item.get("vencimento") or "").strip():
                due = self._parse_date(item["vencimento"])
                if not due:
                    raise ColchaoError(f"Data de vencimento invalida na linha {row_number}.")
                normalized["vencimento"] = due.strftime("%d/%m/%Y")
            if len(normalized) > 2 or normalized.get("status") or normalized.get("vencimento"):
                by_row[row_number] = normalized
        return list(by_row.values())

    def _bucket(self, record: dict[str, Any], profile: str = "alpha") -> str:
        due = self._parse_date(self._due_value(record, profile))
        status = self._status(record)
        today = date.today()
        if not due:
            return "data_invalida"
        if status == "A VENCER" and due == today:
            return "a_vencer_hoje"
        if status == "A VENCER" and due < today:
            return "a_vencer_anterior"
        if status == "VENCIDO" and due == today:
            return "vencida_hoje"
        if status == "VENCIDO" and due < today:
            return "vencida_anterior"
        return status.lower()

    def _ranking(self, rows: list[dict[str, Any]], profile: str = "alpha") -> list[dict[str, Any]]:
        grouped: dict[str, dict[str, Any]] = {}
        for row in rows:
            operador = str(self._operator_value(row, profile) or "Sem operador")
            grouped.setdefault(operador, {"label": operador, "total": 0, "valor": 0.0})
            grouped[operador]["total"] += 1
            grouped[operador]["valor"] += self._money_value(row, profile)
        return sorted(grouped.values(), key=lambda item: item["total"], reverse=True)

    def _copy_row_style(self, sheet, source_row: int, target_row: int, min_col: int, length: int) -> None:
        for offset in range(length):
            source = sheet.cell(source_row, min_col + offset)
            target = sheet.cell(target_row, min_col + offset)
            if source.has_style:
                target._style = copy.copy(source._style)
            target.number_format = source.number_format
            target.font = copy.copy(source.font)
            target.fill = copy.copy(source.fill)
            target.border = copy.copy(source.border)
            target.alignment = copy.copy(source.alignment)
            target.protection = copy.copy(source.protection)

    def _expand_table(self, table, row: int) -> None:
        min_col, min_row, max_col, _max_row = range_boundaries(table.ref)
        table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{row}"

    def _header_lookup(self, headers: list[str], name: str) -> str | None:
        normalized = self._normalize(name)
        return next((header for header in headers if self._normalize(header) == normalized), None)

    def _header_map(self, headers: list[str]) -> dict[str, str]:
        return {self._normalize(header): header for header in headers}

    def _value(self, record: dict[str, Any], name: str) -> Any:
        header = self._header_lookup(list(record.keys()), name)
        return record.get(header or name, "")

    def _first_value(self, record: dict[str, Any], labels: list[str]) -> Any:
        for label in labels:
            value = self._value(record, label)
            if value not in ("", None):
                return value
        return ""

    def _profile_id(self, profile: Any) -> str:
        value = str(profile or "alpha").strip().lower()
        safe = "".join(char for char in value if char.isalnum() or char in {"_", "-"})
        return safe or "alpha"

    def _profile_defaults(self, profile: str) -> dict[str, Any]:
        value = self._profile_id(profile)
        if value in PROFILE_DEFAULTS:
            return PROFILE_DEFAULTS[value]
        return {**GENERIC_PROFILE_DEFAULT, "name": value.replace("_", " ").replace("-", " ").title()}

    def _has_excel_source(self, profile: str) -> bool:
        raw = str(self.get_profile_config(profile).get("excel_path") or "").strip()
        if not raw:
            return False
        return any(Path(candidate).expanduser().is_file() for candidate in self._path_candidates(raw))

    def _required_headers(self, profile: str) -> list[str]:
        profile = self._profile_id(profile)
        custom = self._profile_defaults(profile).get("required_headers") or []
        return custom or REQUIRED_HEADERS

    def _id_value(self, record: dict[str, Any], profile: str = "alpha") -> Any:
        return self._first_value(record, list(self._profile_defaults(profile)["id_labels"]))

    def _client_value(self, record: dict[str, Any], profile: str = "alpha") -> Any:
        return self._first_value(record, list(self._profile_defaults(profile)["client_labels"]))

    def _money_value(self, record: dict[str, Any], profile: str = "alpha") -> float:
        return self._money(self._first_value(record, list(self._profile_defaults(profile)["money_labels"])))

    def _due_value(self, record: dict[str, Any], profile: str = "alpha") -> Any:
        return self._first_value(record, list(self._profile_defaults(profile)["due_labels"]))

    def _operator_value(self, record: dict[str, Any], profile: str = "alpha") -> Any:
        return self._first_value(record, list(self._profile_defaults(profile)["operator_labels"]))

    def _installment_value(self, record: dict[str, Any], profile: str = "alpha") -> Any:
        return self._first_value(record, list(self._profile_defaults(profile)["installment_labels"]))

    def _status(self, record: dict[str, Any]) -> str:
        return self._normalize_status(self._value(record, "STATUS"), strict=False)

    def _normalize_status(self, status: Any, strict: bool = True) -> str:
        value = str(status or "").strip().upper()
        value = " ".join(value.split())
        if value in {"AVENCER", "A_VENCER"}:
            value = "A VENCER"
        if value not in VALID_STATUSES and strict:
            raise ColchaoError("Status invalido.")
        return value

    def _agreement_number(self, value: Any) -> int:
        try:
            return int(float(str(value).replace(",", ".")))
        except ValueError:
            return 0

    def _money(self, value: Any) -> float:
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value or "").replace("R$", "").replace(" ", "").strip()
        if "," in text:
            text = text.replace(".", "").replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return 0.0

    def _parse_date(self, value: Any) -> date | None:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = str(value or "").strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
            try:
                return datetime.strptime(text[:10], fmt).date()
            except ValueError:
                pass
        return None

    def _add_months(self, value: date, months: int) -> date:
        month = value.month - 1 + months
        year = value.year + month // 12
        month = month % 12 + 1
        day = min(value.day, [31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month - 1])
        return date(year, month, day)

    def _normalize(self, value: Any) -> str:
        text = str(value or "").strip().upper()
        without_accents = "".join(char for char in unicodedata.normalize("NFD", text) if unicodedata.category(char) != "Mn")
        return "".join(char for char in without_accents if char.isalnum())

    def _slug(self, value: Any) -> str:
        text = str(value or "").strip().lower()
        without_accents = "".join(char for char in unicodedata.normalize("NFD", text) if unicodedata.category(char) != "Mn")
        return "_".join(part for part in "".join(char if char.isalnum() else " " for char in without_accents).split() if part)

    def _filter_text(self, value: Any) -> str:
        text = str(value or "").strip().lower()
        return "".join(char for char in unicodedata.normalize("NFD", text) if unicodedata.category(char) != "Mn")

    def _cell_value(self, value: Any) -> Any:
        if isinstance(value, datetime):
            return value.strftime("%d/%m/%Y")
        if isinstance(value, date):
            return value.strftime("%d/%m/%Y")
        return "" if value is None else value

    def _log(self, payload: dict[str, Any]) -> None:
        history = self.history()
        history.insert(0, payload)
        self.history_path.write_text(json.dumps(history[:1000], ensure_ascii=False, indent=2, default=str), encoding="utf-8")

    def _read_json(self, path: Path, fallback):
        if not path.exists():
            return fallback
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return fallback
