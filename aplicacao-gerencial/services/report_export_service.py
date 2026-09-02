from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


class ReportExportService:
    DATE_FORMAT = "dd/mm/yyyy"
    ACCOUNTING_FORMAT = '_-"R$"* #,##0.00_-;[Red]-"R$"* #,##0.00_-;_-"R$"* "-"??_-;_-@_-'
    INTEGER_FORMAT = "#,##0"
    PERCENT_FORMAT = "0.00%"
    HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    TABLE_STYLE = "TableStyleMedium2"

    def csv_bytes(self, headers: list[str], rows: list[dict[str, Any]], value_formatter: Callable[[Any], str]) -> bytes:
        output = io.StringIO(newline="")
        writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_MINIMAL)
        writer.writerow(headers)
        for row in rows:
            writer.writerow([value_formatter(row.get(header)) for header in headers])
        return ("\ufeff" + output.getvalue()).encode("utf-8")

    def xlsx_bytes(
        self,
        sheet_name: str,
        headers: list[str],
        rows: list[dict[str, Any]],
        value_formatter: Callable[[Any], str],
    ) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self._safe_sheet_name(sheet_name)
        worksheet.freeze_panes = None
        worksheet.sheet_view.showGridLines = False

        normalized_headers = [str(header or "").strip() or f"COLUNA {index}" for index, header in enumerate(headers, start=1)]
        worksheet.append(normalized_headers)
        for cell in worksheet[1]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")
        worksheet.row_dimensions[1].height = 24

        for row in rows:
            excel_row = worksheet.max_row + 1
            for column_index, header in enumerate(normalized_headers, start=1):
                cell = worksheet.cell(excel_row, column_index)
                raw_value = row.get(headers[column_index - 1])
                value, number_format = self._xlsx_cell_value(header, raw_value, value_formatter)
                cell.value = value
                if number_format:
                    cell.number_format = number_format
                cell.alignment = Alignment(vertical="center")

        if normalized_headers:
            last_column = get_column_letter(len(normalized_headers))
            last_row = max(worksheet.max_row, 1)
            table = Table(displayName=self._safe_table_name(sheet_name), ref=f"A1:{last_column}{last_row}")
            table.tableStyleInfo = TableStyleInfo(
                name=self.TABLE_STYLE,
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            worksheet.add_table(table)
            worksheet.auto_filter.ref = None

        for column_index, header in enumerate(normalized_headers, start=1):
            values = [header, *[value_formatter(row.get(headers[column_index - 1])) for row in rows[:300]]]
            width = min(44, max(11, max((len(str(value or "")) for value in values), default=0) + 2))
            worksheet.column_dimensions[get_column_letter(column_index)].width = width

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _xlsx_cell_value(
        self,
        header: str,
        value: Any,
        value_formatter: Callable[[Any], str],
    ) -> tuple[Any, str | None]:
        if value in (None, ""):
            return None, None
        normalized = self._normalize_header(header)

        if self._is_identifier_header(normalized):
            return value_formatter(value), "@"
        if self._is_date_header(normalized):
            parsed = self._parse_date(value)
            return (parsed, self.DATE_FORMAT) if parsed else (value_formatter(value), None)
        if self._is_percent_header(normalized):
            parsed = self._parse_decimal(value)
            return ((float(parsed) / 100), self.PERCENT_FORMAT) if parsed is not None else (value_formatter(value), None)
        if self._is_money_header(normalized):
            parsed = self._parse_decimal(value)
            return (float(parsed), self.ACCOUNTING_FORMAT) if parsed is not None else (value_formatter(value), None)
        if self._is_integer_header(normalized):
            parsed = self._parse_decimal(value)
            return (int(parsed), self.INTEGER_FORMAT) if parsed is not None else (value_formatter(value), None)
        if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
            return float(value), "#,##0.00"
        return value_formatter(value), None

    @staticmethod
    def _normalize_header(value: Any) -> str:
        import unicodedata

        text = unicodedata.normalize("NFKD", str(value or "").upper())
        return re.sub(r"[^A-Z0-9%]+", " ", "".join(char for char in text if not unicodedata.combining(char))).strip()

    @classmethod
    def _is_date_header(cls, header: str) -> bool:
        return (
            header.startswith("DATA")
            or header.startswith("DT ")
            or " VENCIMENTO" in f" {header}"
            or header in {"MES", "COMPETENCIA", "ATUALIZACAO", "CRIADO EM"}
        )

    @staticmethod
    def _is_identifier_header(header: str) -> bool:
        return any(
            token in header
            for token in ("NPJ", "CPF", "CNPJ", "DEBIT ID", "SUITID", "PROCESSO", "CONTRATO", "GECOR", "PJ", "PK")
        )

    @staticmethod
    def _is_money_header(header: str) -> bool:
        return any(
            token in header
            for token in ("VALOR", "HONORARIO", "H O", "CASH", "ENTRADA", "OFERTA", "ACEITO", "PAGO")
        )

    @staticmethod
    def _is_percent_header(header: str) -> bool:
        return header in {"%", "% H O", "PERCENTUAL", "PERCENTUAL H O"} or "PERCENTUAL" in header

    @staticmethod
    def _is_integer_header(header: str) -> bool:
        return any(token in header for token in ("DIAS", "QUANTIDADE", "QTD", "PARCELAS"))

    @staticmethod
    def _parse_date(value: Any) -> date | datetime | None:
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return value
        text = str(value or "").strip()
        if not text:
            return None
        if re.fullmatch(r"\d{4}-\d{2}", text):
            try:
                return datetime.strptime(f"{text}-01", "%Y-%m-%d").date()
            except ValueError:
                return None
        for pattern in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y"):
            try:
                return datetime.strptime(text[:19] if "%S" in pattern else text[:10], pattern).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def _parse_decimal(value: Any) -> Decimal | None:
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return Decimal(str(value))
        text = str(value or "").strip().replace("R$", "").replace("\u00a0", "").replace(" ", "")
        if not text:
            return None
        if "," in text:
            text = text.replace(".", "").replace(",", ".")
        try:
            return Decimal(text)
        except InvalidOperation:
            return None

    @staticmethod
    def _safe_sheet_name(value: str) -> str:
        clean = re.sub(r"[\[\]:*?/\\]", " ", str(value or "Relatorio")).strip()
        return (clean or "Relatorio")[:31]

    @classmethod
    def _safe_table_name(cls, value: str) -> str:
        clean = cls._normalize_header(value).title().replace(" ", "")
        clean = re.sub(r"[^A-Za-z0-9_]", "", clean)
        if not clean or clean[0].isdigit():
            clean = f"Tabela{clean}"
        return f"{clean[:220]}Dados"
