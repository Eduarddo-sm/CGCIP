from __future__ import annotations

import csv
import io
import json
import shutil
import threading
import time
import unicodedata
import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

from services.negocial_service import NegocialService


class ParecerError(RuntimeError):
    pass


DEFAULT_CONFIG = {
    "excel_path": "data/demo/parecer-demo.xlsx",
    "main_sheet": "PARECER GAMMA",
    "control_sheet": "P.CONCLUÍDOS",
    "pk_column": "PK",
    "solicitado_column": "SOLICITADO?",
    "auto_refresh_minutes": 10,
}


class ParecerService:
    NEGOCIAL_PREFIX = "NEGOCIAL:"

    def __init__(self, data_dir: Path, negocial: NegocialService | None = None) -> None:
        self.data_dir = data_dir
        self.negocial = negocial
        self.config_path = data_dir / "parecer_config.json"
        self.history_path = data_dir / "parecer_history.json"
        self.backup_dir = data_dir / "backups"
        self._excel_lock = threading.Lock()
        self._cache_lock = threading.Lock()
        self._records_cache: dict[str, Any] | None = None
        self.backup_dir.mkdir(exist_ok=True)
        if not self.config_path.exists():
            self.save_config(DEFAULT_CONFIG)

    def get_config(self) -> dict[str, Any]:
        config = DEFAULT_CONFIG | self._read_json(self.config_path, {})
        return config

    def save_config(self, payload: dict[str, Any]) -> dict[str, Any]:
        config = self.get_config() if self.config_path.exists() else DEFAULT_CONFIG.copy()
        for key in DEFAULT_CONFIG:
            if key in payload:
                config[key] = payload[key]
        config["auto_refresh_minutes"] = int(config.get("auto_refresh_minutes") or 10)
        self.config_path.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
        self.clear_cache()
        return config

    def read_records(self) -> list[dict[str, Any]]:
        signature = self._cache_signature()
        with self._cache_lock:
            if self._records_cache and self._records_cache.get("signature") == signature:
                return copy.deepcopy(self._records_cache["records"])
        records: list[dict[str, Any]] = []
        try:
            records.extend(self._read_records_from_excel())
        except ParecerError as exc:
            self._log("", "read_excel_pareceres", "sistema", "erro", str(exc))
        records.extend(self._read_records_from_negocial())
        for record in records:
            record.setdefault("DATA SOLICITADO", "")
            record.setdefault("DATA APROVADO/REPROVADO", "")
        with self._cache_lock:
            self._records_cache = {
                "signature": signature,
                "records": copy.deepcopy(records),
                "updated_at": datetime.now().isoformat(timespec="seconds"),
            }
        return records

    def refresh_cache(self) -> dict[str, Any]:
        self.clear_cache()
        records = self.read_records()
        pendentes = self.read_pendentes()
        return {
            "ok": True,
            "records": len(records),
            "pendentes": len(pendentes),
            "updated_at": datetime.now().isoformat(timespec="seconds"),
        }

    def clear_cache(self) -> None:
        with self._cache_lock:
            self._records_cache = None

    def _read_records_from_excel(self) -> list[dict[str, Any]]:
        config = self.get_config()
        workbook = self._load_workbook(data_only=True)
        try:
            sheet_name = config["main_sheet"]
            if sheet_name not in workbook.sheetnames:
                raise ParecerError(f"Aba principal '{sheet_name}' nao encontrada.")
            sheet = workbook[sheet_name]
            table = self._find_table(sheet, [config["solicitado_column"]])
            headers, header_row, min_col, max_row = self._headers(sheet, table.ref if table else None)
            solicitado = self._header_lookup(headers, config["solicitado_column"])
            records: list[dict[str, Any]] = []
            for row_index in range(header_row + 1, max_row + 1):
                values = [self._cell_value(sheet.cell(row_index, min_col + col_index).value) for col_index in range(len(headers))]
                if not any(value not in ("", None) for value in values):
                    continue
                record = {header: values[index] for index, header in enumerate(headers)}
                record["__row_number"] = row_index
                if solicitado and solicitado not in record:
                    record[config["solicitado_column"]] = ""
                records.append(record)
            return records
        finally:
            workbook.close()

    def read_pendentes(self) -> list[dict[str, Any]]:
        config = self.get_config()
        records = [record for record in self.read_records() if str(record.get("__source") or "excel") != "sistema"]
        solicitado_header = self._header_lookup_from_records(records, config["solicitado_column"])
        pendentes = [record for record in records if self._is_not_requested(record.get(solicitado_header or config["solicitado_column"]))]
        pendentes.extend(self._read_pendentes_from_negocial())
        return pendentes

    def read_aprovacao_pendente(self) -> list[dict[str, Any]]:
        if not self.negocial:
            return []
        try:
            return self.negocial.read_parecer_approval_pending()
        except Exception as exc:
            self._log("", "read_pareceres_aprovacao", "sistema", "erro", str(exc))
            return []

    def read_aprovacao_historico(self) -> list[dict[str, Any]]:
        if not self.negocial:
            return []
        try:
            return self.negocial.read_parecer_approval_history()
        except Exception as exc:
            self._log("", "read_pareceres_aprovacao_historico", "sistema", "erro", str(exc))
            return []

    def dashboard(self) -> dict[str, Any]:
        records = self.read_records()
        config = self.get_config()
        excel_records = [record for record in records if str(record.get("__source") or "excel") != "sistema"]
        sistema_records = [record for record in records if str(record.get("__source") or "") == "sistema"]
        solicitado_header = self._header_lookup_from_records(excel_records, config["solicitado_column"])
        pendentes = [record for record in excel_records if self._is_not_requested(record.get(solicitado_header or config["solicitado_column"]))]
        pendentes.extend([
            record for record in sistema_records
            if str(record.get("STATUS") or "").upper() == "PENDENTE"
            and str(record.get("APROVACAO") or "").upper() == "APROVADO"
        ])
        solicitados = [record for record in excel_records if self._normalize(record.get(solicitado_header or config["solicitado_column"])) == "SIM"]
        solicitados.extend([record for record in sistema_records if str(record.get("STATUS") or "").upper() == "SOLICITADO"])
        aguardando_aprovacao = [
            record for record in sistema_records
            if str(record.get("STATUS") or "").upper() == "PENDENTE"
            and str(record.get("APROVACAO") or "PENDENTE").upper() == "PENDENTE"
        ]
        aprovados = [record for record in sistema_records if str(record.get("APROVACAO") or "").upper() == "APROVADO"]
        reprovados = [
            record for record in sistema_records
            if str(record.get("APROVACAO") or "").upper() == "REPROVADO"
            or str(record.get("STATUS") or "").upper() == "CANCELADO"
        ]
        fila_atencao = [
            self._dashboard_record(record, "Aguardando aprovacao", "aprovacao")
            for record in aguardando_aprovacao
        ]
        fila_atencao.extend([
            self._dashboard_record(record, "Pronto para solicitar", "pendentes")
            for record in pendentes
        ])
        fila_atencao.sort(key=lambda item: (0 if item["target"] == "aprovacao" else 1, item["sort_at"]))
        atividade_recente = sorted(
            [self._dashboard_record(record, self._dashboard_action(record, solicitado_header, config), "") for record in records],
            key=lambda item: item["sort_at"],
            reverse=True,
        )[:10]
        return {
            "total": len(records),
            "pendentes": len(pendentes),
            "solicitados": len(solicitados),
            "aguardando_aprovacao": len(aguardando_aprovacao),
            "aprovados": len(aprovados),
            "reprovados": len(reprovados),
            "por_negociador": self._count_by(pendentes, ["OPERADOR", "NEGOCIADOR", "Negociador", "responsavel", "RESPONSAVEL"]),
            "por_data": self._count_by_date(pendentes, ["DATA", "Data", "data"]),
            "tendencia": self._count_by_date(records, ["DATA", "Data", "data"]),
            "fila_atencao": [{key: value for key, value in item.items() if key != "sort_at"} for item in fila_atencao[:10]],
            "atividade_recente": [{key: value for key, value in item.items() if key != "sort_at"} for item in atividade_recente],
            "updated_at": datetime.now().isoformat(timespec="seconds"),
        }

    def _dashboard_record(self, record: dict[str, Any], action: str, target: str) -> dict[str, Any]:
        raw_date = self._first_value(record, ["__updated_at", "__created_at", "DATA", "Data", "data"])
        return {
            "pk": str(self._first_value(record, ["PK", "NPJ"]) or ""),
            "cliente": str(self._first_value(record, ["NOME CLIENTE", "CLIENTE", "NOME", "NOME DO CLIENTE"]) or "Cliente nao identificado"),
            "negociador": str(self._first_value(record, ["OPERADOR", "NEGOCIADOR", "RESPONSAVEL", "RESPONSÁVEL", "USUARIO"]) or "Nao informado"),
            "motivo": str(self._first_value(record, ["MOTIVO", "MOTIVO PARECER", "TIPO MOTIVO"]) or "Motivo nao informado"),
            "acao": action,
            "target": target,
            "data": self._date_label(raw_date),
            "sort_at": self._dashboard_timestamp(raw_date),
        }

    def _dashboard_action(self, record: dict[str, Any], solicitado_header: str | None, config: dict[str, Any]) -> str:
        approval = str(record.get("APROVACAO") or "").upper()
        status = str(record.get("STATUS") or "").upper()
        solicitado = self._normalize(record.get(solicitado_header or config["solicitado_column"]))
        if approval == "REPROVADO" or status == "CANCELADO":
            return "Parecer reprovado"
        if status == "SOLICITADO" or solicitado == "SIM":
            return "Parecer solicitado"
        if approval == "APROVADO":
            return "Parecer aprovado"
        if approval == "PENDENTE":
            return "Recebido para aprovacao"
        return "Parecer atualizado"

    def _dashboard_timestamp(self, value: Any) -> float:
        if isinstance(value, datetime):
            return value.timestamp()
        text = str(value or "").strip()
        if not text:
            return 0.0
        try:
            return datetime.fromisoformat(text.replace("Z", "+00:00")).timestamp()
        except (TypeError, ValueError):
            return 0.0

    def relatorio_csv(self, carteira: str = "") -> tuple[str, bytes]:
        rows = self.read_records()
        wallet = str(carteira or "").strip().upper()
        if wallet:
            rows = [
                row for row in rows
                if str(row.get("CARTEIRA") or row.get("carteira") or "GAMMA").strip().upper() == wallet
            ]
        headers = self._report_headers(rows)
        output = io.StringIO(newline="")
        writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_MINIMAL)
        writer.writerow(headers)
        for row in rows:
            writer.writerow([self._csv_value(self._report_value(row, header)) for header in headers])
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        suffix = f"_{wallet.lower()}" if wallet else ""
        return f"relatorio_pareceres{suffix}_{stamp}.csv", ("\ufeff" + output.getvalue()).encode("utf-8")

    def marcar_solicitado(self, pk: str, user: str) -> dict[str, Any]:
        pk = str(pk or "").strip()
        if not pk:
            raise ParecerError("PK nao informada.")
        if self._is_negocial_pk(pk):
            return self._marcar_solicitado_negocial(pk, user)
        started = time.perf_counter()
        with self._excel_lock:
            backup = self.create_backup()
            config = self.get_config()
            try:
                result = self._marcar_solicitado_excel_com(pk, config)
            except Exception as exc:
                detail = str(exc) or repr(exc)
                self._log(pk, "marcar_solicitado", user, "erro", detail)
                raise
            self.clear_cache()
        elapsed_ms = int((time.perf_counter() - started) * 1000)
        status = "duplicado" if result.get("duplicated") else "ok"
        detail = "PK ja existia em P.CONCLUIDOS" if result.get("duplicated") else f"Inserido na linha {result.get('row')} em {elapsed_ms} ms."
        self._log(pk, "marcar_solicitado", user, status, detail)
        return {"ok": True, "pk": pk, "backup": str(backup), "elapsed_ms": elapsed_ms, **result}

    def marcar_varios(self, pks: list[str], user: str) -> dict[str, Any]:
        normalized_pks = [str(pk or "").strip() for pk in pks]
        normalized_pks = [pk for pk in normalized_pks if pk]
        if not normalized_pks:
            raise ParecerError("Nenhuma PK informada.")
        system_pks = [pk for pk in normalized_pks if self._is_negocial_pk(pk)]
        excel_pks = [pk for pk in normalized_pks if not self._is_negocial_pk(pk)]
        if system_pks and not excel_pks:
            results = [self._marcar_solicitado_negocial(pk, user) for pk in system_pks]
            return {"ok": True, "results": results, "elapsed_ms": 0}
        if system_pks and excel_pks:
            excel_result = self.marcar_varios(excel_pks, user)
            system_results = [self._marcar_solicitado_negocial(pk, user) for pk in system_pks]
            return {
                "ok": True,
                "backup": excel_result.get("backup"),
                "elapsed_ms": excel_result.get("elapsed_ms", 0),
                "results": list(excel_result.get("results", [])) + system_results,
            }
        started = time.perf_counter()
        with self._excel_lock:
            backup = self.create_backup()
            config = self.get_config()
            try:
                payload = self._marcar_varios_excel_com(normalized_pks, config)
            except Exception as exc:
                detail = str(exc) or repr(exc)
                self._log(", ".join(normalized_pks), "marcar_varios", user, "erro", detail)
                raise
            self.clear_cache()
        elapsed_ms = int((time.perf_counter() - started) * 1000)
        for result in payload["results"]:
            status = "duplicado" if result.get("duplicated") else "ok"
            detail = "PK ja existia em P.CONCLUIDOS" if result.get("duplicated") else f"Inserido na linha {result.get('row')} em lote."
            self._log(result["pk"], "marcar_solicitado", user, status, detail)
        return {"ok": True, "backup": str(backup), "elapsed_ms": elapsed_ms, **payload}

    def refresh_powerquery(self, user: str = "sistema") -> dict[str, Any]:
        path = self._excel_path()
        try:
            import pythoncom  # type: ignore
            import win32com.client  # type: ignore
        except ImportError:
            message = "Atualizacao Power Query requer Microsoft Excel e pywin32/win32com instalados neste Python."
            self._log("", "powerquery_refresh", user, "indisponivel", message)
            return {"ok": False, "message": message}
        excel = None
        workbook = None
        close_workbook = False
        quit_excel = False
        pythoncom.CoInitialize()
        try:
            excel, workbook, close_workbook, quit_excel = self._open_workbook_com(win32com.client, path, read_only=False)
            self._refresh_open_workbook(excel, workbook)
            self._com_call(lambda: workbook.Save())
            self.clear_cache()
            if close_workbook:
                self._com_call(lambda: workbook.Close(False))
                workbook = None
            records = self.read_records()
            solicitado_header = self._header_lookup_from_records(records, self.get_config()["solicitado_column"])
            pendentes = len([record for record in records if self._is_not_requested(record.get(solicitado_header or self.get_config()["solicitado_column"]))])
            total = len(records)
            detail = f"Atualizar Tudo executado e planilha salva. {pendentes} pendentes de {total} registros."
            self._log("", "powerquery_refresh", user, "ok", detail)
            return {"ok": True, "message": f"Atualizar Tudo executado. {pendentes} pendentes encontrados.", "total": total, "pendentes": pendentes}
        except Exception as exc:  # pragma: no cover - depends on local Excel COM
            self._log("", "powerquery_refresh", user, "erro", str(exc))
            raise ParecerError(f"Falha ao atualizar Power Query: {exc}") from exc
        finally:
            if workbook is not None and close_workbook:
                try:
                    workbook.Close(False)
                except Exception:
                    pass
            if excel is not None and quit_excel:
                excel.Quit()
            pythoncom.CoUninitialize()

    def history(self) -> list[dict[str, Any]]:
        return self._read_json(self.history_path, [])

    def create_backup(self) -> Path:
        path = self._excel_path()
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        target = self.backup_dir / f"{path.stem}_{stamp}{path.suffix}"
        shutil.copy2(path, target)
        return target

    def _report_headers(self, rows: list[dict[str, Any]]) -> list[str]:
        preferred = [
            "PK",
            "DATA",
            "NPJ",
            "NOME CLIENTE",
            "CLIENTE",
            "NEGOCIADOR",
            "MOTIVO",
            "DESCRIÇÃO",
            "DESCRICAO",
            "SOLICITADO?",
            "STATUS",
        ]
        negotiator_aliases = {
            self._normalize(alias)
            for alias in [
                "NEGOCIADOR",
                "OPERADOR",
                "RESPONSAVEL",
                "RESPONSÁVEL",
                "SOLICITANTE",
                "USUARIO",
                "USUÁRIO",
            ]
        }
        seen: set[str] = set()
        seen_normalized: set[str] = set()
        headers: list[str] = []
        for header in preferred:
            if header == "NEGOCIADOR":
                seen.add(header)
                seen_normalized.update(negotiator_aliases)
                headers.append(header)
                continue
            key = self._matching_key(rows, header)
            normalized_key = self._normalize(key)
            if key and key not in seen and normalized_key not in seen_normalized:
                seen.add(key)
                seen_normalized.add(normalized_key)
                headers.append(key)
        for row in rows:
            for key in row:
                normalized_key = self._normalize(key)
                if str(key).startswith("__") or key in seen or normalized_key in seen_normalized:
                    continue
                seen.add(key)
                seen_normalized.add(normalized_key)
                headers.append(key)
        return headers

    def _report_value(self, row: dict[str, Any], header: str) -> Any:
        if self._normalize(header) == "NEGOCIADOR":
            return self._first_value(
                row,
                [
                    "NEGOCIADOR",
                    "OPERADOR",
                    "Negociador",
                    "Operador",
                    "RESPONSAVEL",
                    "RESPONSÁVEL",
                    "responsavel",
                    "SOLICITANTE",
                    "USUARIO",
                    "USUÁRIO",
                    "usuario",
                    "operador",
                ],
            )
        return row.get(header, "")

    def _first_value(self, row: dict[str, Any], keys: list[str]) -> Any:
        normalized_map = {self._normalize(key): value for key, value in row.items()}
        for key in keys:
            value = row.get(key)
            if value not in (None, ""):
                return value
            value = normalized_map.get(self._normalize(key))
            if value not in (None, ""):
                return value
        return ""

    def _matching_key(self, rows: list[dict[str, Any]], header: str) -> str:
        wanted = self._normalize(header)
        for row in rows:
            for key in row:
                if self._normalize(key) == wanted:
                    return key
        return ""

    def _csv_value(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%d/%m/%Y %H:%M:%S")
        return str(value).replace("\r", " ").replace("\n", " ").strip()

    def excel_path(self) -> Path:
        return self._excel_path()

    def _read_records_from_negocial(self) -> list[dict[str, Any]]:
        if not self.negocial:
            return []
        try:
            return self.negocial.read_parecer_records()
        except Exception as exc:
            self._log("", "read_negocial_pareceres", "sistema", "erro", str(exc))
            return []

    def _read_pendentes_from_negocial(self) -> list[dict[str, Any]]:
        if not self.negocial:
            return []
        try:
            return self.negocial.read_parecer_pendentes()
        except Exception as exc:
            self._log("", "read_negocial_pareceres_pendentes", "sistema", "erro", str(exc))
            return []

    def _marcar_solicitado_negocial(self, pk: str, user: str) -> dict[str, Any]:
        if not self.negocial:
            raise ParecerError("Integracao com pareceres do sistema negocial nao configurada.")
        parecer_id = self._negocial_id_from_pk(pk)
        started = time.perf_counter()
        try:
            result = self.negocial.marcar_parecer_solicitado(parecer_id, user)
        except Exception as exc:
            self._log(pk, "marcar_solicitado_negocial", user, "erro", str(exc))
            raise ParecerError(str(exc)) from exc
        self.clear_cache()
        elapsed_ms = int((time.perf_counter() - started) * 1000)
        status = "duplicado" if result.get("duplicated") else "ok"
        detail = "Parecer negocial ja estava solicitado" if result.get("duplicated") else f"Status alterado no sistema negocial em {elapsed_ms} ms."
        self._log(pk, "marcar_solicitado_negocial", user, status, detail)
        return {"ok": True, "pk": pk, "elapsed_ms": elapsed_ms, **result}

    def aprovar_negocial(self, pk: str, reason: str, descricao: str, user: str) -> dict[str, Any]:
        if not self.negocial:
            raise ParecerError("Integracao com pareceres do sistema negocial nao configurada.")
        parecer_id = self._negocial_id_from_pk(pk)
        try:
            result = self.negocial.aprovar_parecer(parecer_id, reason, descricao, user)
        except Exception as exc:
            self._log(pk, "aprovar_parecer_negocial", user, "erro", str(exc))
            raise ParecerError(str(exc)) from exc
        self.clear_cache()
        self._log(pk, "aprovar_parecer_negocial", user, "ok", str(reason or "")[:600])
        return result

    def reprovar_negocial(self, pk: str, reason: str, descricao: str, user: str) -> dict[str, Any]:
        if not self.negocial:
            raise ParecerError("Integracao com pareceres do sistema negocial nao configurada.")
        parecer_id = self._negocial_id_from_pk(pk)
        try:
            result = self.negocial.reprovar_parecer(parecer_id, reason, descricao, user)
        except Exception as exc:
            self._log(pk, "reprovar_parecer_negocial", user, "erro", str(exc))
            raise ParecerError(str(exc)) from exc
        self.clear_cache()
        self._log(pk, "reprovar_parecer_negocial", user, "ok", str(reason or "")[:600])
        return result

    def _is_negocial_pk(self, pk: str) -> bool:
        return str(pk or "").upper().startswith(self.NEGOCIAL_PREFIX)

    def _negocial_id_from_pk(self, pk: str) -> int:
        try:
            return int(str(pk).split(":", 1)[1])
        except (IndexError, TypeError, ValueError) as exc:
            raise ParecerError(f"PK negocial invalida: {pk}") from exc

    def _cache_signature(self) -> tuple[str, float, str, str, str, str]:
        config = self.get_config()
        try:
            path = self._excel_path()
            path_value = str(path.resolve())
            mtime = path.stat().st_mtime
            error = ""
        except ParecerError as exc:
            path_value = str(config.get("excel_path", ""))
            mtime = 0
            error = str(exc)
        try:
            negocial_marker = self.negocial.parecer_marker() if self.negocial else ""
        except Exception as exc:
            negocial_marker = f"erro:{exc}"
        return (
            path_value,
            mtime,
            str(config.get("main_sheet", "")),
            str(config.get("solicitado_column", "")),
            error,
            negocial_marker,
        )

    def _excel_path(self) -> Path:
        raw_path = Path(str(self.get_config()["excel_path"]))
        path = raw_path if raw_path.is_absolute() else (self.data_dir.parent / raw_path)
        if not path.exists():
            raise ParecerError(f"Arquivo Excel nao encontrado: {path}")
        return path

    def _load_workbook(self, data_only: bool):
        try:
            return load_workbook(self._excel_path(), data_only=data_only, keep_vba=True)
        except PermissionError as exc:
            raise ParecerError("Nao foi possivel acessar a planilha. Verifique se o arquivo esta bloqueado.") from exc
        except Exception as exc:
            raise ParecerError(f"Falha ao abrir planilha: {exc}") from exc

    def _headers(self, sheet, table_ref: str | None = None) -> tuple[list[str], int, int, int]:
        if table_ref:
            min_col, min_row, max_col, max_row = range_boundaries(table_ref)
            values = [self._cell_value(sheet.cell(min_row, col).value) for col in range(min_col, max_col + 1)]
            headers = [
                str(value).strip() if value not in ("", None) else f"COLUNA {index}"
                for index, value in enumerate(values, start=1)
            ]
            return headers, min_row, min_col, max_row
        for row_index in range(1, min(sheet.max_row, 25) + 1):
            values = [self._cell_value(sheet.cell(row_index, col).value) for col in range(1, sheet.max_column + 1)]
            filled = [value for value in values if value not in ("", None)]
            if len(filled) >= 2:
                headers = []
                for index, value in enumerate(values, start=1):
                    header = str(value).strip() if value not in ("", None) else f"COLUNA {index}"
                    headers.append(header)
                return headers, row_index, 1, sheet.max_row
        raise ParecerError("Nao foi possivel identificar cabecalhos na aba principal.")

    def _existing_control_pks(self, sheet, column: int, start_row: int, end_row: int) -> set[str]:
        values = set()
        for row in range(start_row, end_row + 1):
            value = self._cell_value(sheet.cell(row, column).value)
            if value:
                values.add(str(value).strip())
        return values

    def _next_empty_row(self, sheet, column: int, start_row: int, end_row: int) -> int:
        for row in range(start_row, end_row + 1):
            if not self._cell_value(sheet.cell(row, column).value):
                return row
        return end_row + 1

    def _find_table(self, sheet, required_headers: list[str] | None = None):
        required = [self._normalize(header) for header in (required_headers or [])]
        fallback = None
        for table in self._tables(sheet):
            fallback = fallback or table
            headers, _, _, _ = self._headers(sheet, table.ref)
            normalized_headers = {self._normalize(header) for header in headers}
            if all(header in normalized_headers for header in required):
                return table
        return fallback

    def _first_table(self, sheet):
        tables = self._tables(sheet)
        return tables[0] if tables else None

    def _tables(self, sheet) -> list:
        tables = getattr(sheet, "tables", {})
        if not tables:
            return []
        result = []
        for table in tables.values():
            if hasattr(table, "ref"):
                result.append(table)
            if isinstance(table, str):
                result.append(tables[table])
        return result

    def _expand_table(self, table, row: int) -> None:
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        if row > max_row:
            table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{row}"

    def _marcar_solicitado_excel_com(self, pk: str, config: dict[str, Any]) -> dict[str, Any]:
        payload = self._marcar_varios_excel_com([pk], config)
        return payload["results"][0]

    def _marcar_varios_excel_com(self, pks: list[str], config: dict[str, Any]) -> dict[str, Any]:
        path = self._excel_path()
        try:
            import pythoncom  # type: ignore
            import win32com.client  # type: ignore
        except ImportError as exc:
            raise ParecerError("Para gravar nesta planilha com Power Query, o pywin32 precisa estar instalado.") from exc

        excel = None
        workbook = None
        close_workbook = False
        quit_excel = False
        pythoncom.CoInitialize()
        try:
            excel, workbook, close_workbook, quit_excel = self._open_workbook_com(win32com.client, path, read_only=False)
            sheet = self._com_call(lambda: workbook.Worksheets(config["control_sheet"]))
            table = self._find_com_table(sheet, ["NPJ", "SOLICITADO"])
            if table is None:
                raise ParecerError(f"Nenhuma tabela formatada com NPJ/SOLICITADO foi encontrada na aba '{config['control_sheet']}'.")
            npj_index = self._com_table_column_index(table, "NPJ")
            solicitado_index = self._com_table_column_index(table, "SOLICITADO")
            if not npj_index:
                raise ParecerError("A coluna NPJ nao foi encontrada na tabela de P.CONCLUIDOS.")
            if not solicitado_index:
                raise ParecerError("A coluna SOLICITADO nao foi encontrada na tabela de P.CONCLUIDOS.")
            results = []
            changed = False
            for pk in pks:
                existing_row = self._com_find_value_in_table_column(table, npj_index, pk)
                if existing_row:
                    results.append({"ok": True, "pk": pk, "duplicated": True, "row": existing_row})
                    continue
                list_row = self._com_call(lambda: table.ListRows.Add())
                npj_cell = list_row.Range.Cells(1, npj_index)
                npj_cell.NumberFormat = "@"
                npj_cell.Value = str(pk)
                list_row.Range.Cells(1, solicitado_index).Value = "SIM"
                row_number = int(list_row.Range.Row)
                results.append({"ok": True, "pk": pk, "duplicated": False, "row": row_number})
                changed = True
            if changed:
                self._calculate_request_status(excel, workbook, config)
                self._com_call(lambda: workbook.Save())
            if close_workbook:
                self._com_call(lambda: workbook.Close(False))
                workbook = None
            return {"results": results}
        except Exception as exc:
            if workbook is not None and close_workbook:
                try:
                    workbook.Close(False)
                except Exception:
                    pass
            if isinstance(exc, ParecerError):
                raise
            raise ParecerError(f"Falha ao gravar no Excel via COM: {exc}") from exc
        finally:
            if excel is not None and quit_excel:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()

    def _find_com_table(self, sheet, required_headers: list[str]):
        required = {self._normalize(header) for header in required_headers}
        fallback = None
        for table in self._com_iter(sheet.ListObjects):
            fallback = fallback or table
            headers = {self._normalize(column.Name) for column in self._com_iter(table.ListColumns)}
            if required.issubset(headers):
                return table
        return fallback

    def _com_table_column_index(self, table, header: str) -> int | None:
        normalized = self._normalize(header)
        for column in self._com_iter(table.ListColumns):
            if self._normalize(column.Name) == normalized:
                return int(column.Index)
        return None

    def _com_find_value_in_table_column(self, table, column_index: int, value: str) -> int | None:
        data_body = table.DataBodyRange
        if data_body is None:
            return None
        expected = str(value).strip()
        column_range = data_body.Columns(column_index)
        try:
            found = column_range.Find(What=expected, LookIn=-4163, LookAt=1, SearchOrder=1, SearchDirection=1, MatchCase=False)
            if found is not None:
                return int(found.Row)
        except Exception:
            pass
        try:
            values = column_range.Value
            for row_index, current in enumerate(self._flatten_com_column(values), start=1):
                if str(current or "").strip() == expected:
                    return int(data_body.Cells(row_index, column_index).Row)
        except Exception:
            pass
        return None

    def _flatten_com_column(self, values) -> list[Any]:
        if isinstance(values, tuple):
            result = []
            for item in values:
                if isinstance(item, tuple):
                    result.append(item[0] if item else "")
                else:
                    result.append(item)
            return result
        return [values]

    def _open_workbook_com(self, win32, path: Path, read_only: bool = False):
        target = self._normalize_path(str(path))
        try:
            excel = self._com_call(lambda: win32.GetActiveObject("Excel.Application"), timeout_seconds=3)
            for workbook in self._com_iter(excel.Workbooks):
                try:
                    if self._normalize_path(str(workbook.FullName)) == target:
                        workbook.Application.DisplayAlerts = False
                        if not read_only and bool(workbook.ReadOnly):
                            raise ParecerError("A planilha esta aberta como somente leitura. Feche e abra com permissao de edicao.")
                        return workbook.Application, workbook, False, False
                except ParecerError:
                    raise
                except Exception:
                    pass
        except ParecerError:
            raise
        except Exception:
            pass

        attached = self._get_workbook_object_by_path(win32, path)
        if attached is not None:
            excel, workbook = attached
            excel.DisplayAlerts = False
            if not read_only and bool(workbook.ReadOnly):
                raise ParecerError("A planilha esta aberta como somente leitura. Feche e abra com permissao de edicao.")
            return excel, workbook, False, False

        excel = self._com_call(lambda: win32.DispatchEx("Excel.Application"))
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        workbook = self._com_call(lambda: excel.Workbooks.Open(str(path), UpdateLinks=0, ReadOnly=read_only))
        if not read_only and bool(workbook.ReadOnly):
            try:
                workbook.Close(False)
            finally:
                excel.Quit()
            raise ParecerError("A planilha foi aberta como somente leitura. Verifique se ela ja esta aberta ou bloqueada na rede.")
        return excel, workbook, True, True

    def _get_workbook_object_by_path(self, win32, path: Path):
        try:
            workbook = self._com_call(lambda: win32.GetObject(str(path)), timeout_seconds=8)
            if self._normalize_path(str(workbook.FullName)) != self._normalize_path(str(path)):
                return None
            return workbook.Application, workbook
        except Exception:
            return None

    def _normalize_path(self, value: str) -> str:
        return value.replace("/", "\\").rstrip("\\").casefold()

    def _com_call(self, action, timeout_seconds: int = 30):
        deadline = time.time() + timeout_seconds
        last_exc = None
        while time.time() < deadline:
            try:
                return action()
            except Exception as exc:
                last_exc = exc
                text = str(exc)
                if "-2147418111" not in text and "rejeitada" not in text.lower() and "rejected" not in text.lower():
                    raise
                time.sleep(0.35)
        raise last_exc

    def _disable_background_refresh(self, workbook) -> None:
        for connection in self._com_iter(workbook.Connections):
            for attr in ("OLEDBConnection", "ODBCConnection"):
                try:
                    getattr(connection, attr).BackgroundQuery = False
                except Exception:
                    pass
        for sheet in self._com_iter(workbook.Worksheets):
            for query_table in self._com_iter(sheet.QueryTables):
                try:
                    query_table.BackgroundQuery = False
                except Exception:
                    pass
            for list_object in self._com_iter(sheet.ListObjects):
                try:
                    if list_object.QueryTable is not None:
                        list_object.QueryTable.BackgroundQuery = False
                except Exception:
                    pass

    def _wait_for_excel_refresh(self, excel, workbook, timeout_seconds: int = 300) -> None:
        deadline = time.time() + timeout_seconds
        while time.time() < deadline:
            excel.CalculateUntilAsyncQueriesDone()
            if not self._has_refreshing_connection(workbook):
                return
            time.sleep(0.5)
        raise ParecerError("Atualizar Tudo demorou demais para finalizar.")

    def _has_refreshing_connection(self, workbook) -> bool:
        for connection in self._com_iter(workbook.Connections):
            for attr in ("OLEDBConnection", "ODBCConnection"):
                try:
                    if getattr(connection, attr).Refreshing:
                        return True
                except Exception:
                    pass
        for sheet in self._com_iter(workbook.Worksheets):
            for query_table in self._com_iter(sheet.QueryTables):
                try:
                    if query_table.Refreshing:
                        return True
                except Exception:
                    pass
            for list_object in self._com_iter(sheet.ListObjects):
                try:
                    if list_object.QueryTable is not None and list_object.QueryTable.Refreshing:
                        return True
                except Exception:
                    pass
        return False

    def _refresh_open_workbook(self, excel, workbook) -> None:
        self._disable_background_refresh(workbook)
        self._com_call(lambda: workbook.RefreshAll())
        self._wait_for_excel_refresh(excel, workbook)
        self._calculate_open_workbook(excel)

    def _calculate_request_status(self, excel, workbook, config: dict[str, Any]) -> None:
        for sheet_name in (config["control_sheet"], config["main_sheet"]):
            try:
                sheet = self._com_call(lambda name=sheet_name: workbook.Worksheets(name), timeout_seconds=5)
                self._com_call(lambda target=sheet: target.Calculate(), timeout_seconds=20)
            except Exception:
                pass
        try:
            self._com_call(lambda: excel.CalculateUntilAsyncQueriesDone(), timeout_seconds=20)
        except Exception:
            pass

    def _calculate_open_workbook(self, excel) -> None:
        self._com_call(lambda: excel.CalculateFullRebuild())
        self._com_call(lambda: excel.CalculateUntilAsyncQueriesDone())

    def _com_iter(self, collection):
        try:
            count = int(collection.Count)
        except Exception:
            return []
        return [collection.Item(index) for index in range(1, count + 1)]

    def _log(self, pk: str, action: str, user: str, result: str, detail: str) -> None:
        history = self.history()
        history.insert(0, {
            "data_hora": datetime.now().isoformat(timespec="seconds"),
            "pk": pk,
            "acao": action,
            "usuario": user,
            "resultado": result,
            "detalhe": detail,
        })
        self.history_path.write_text(json.dumps(history[:1000], ensure_ascii=False, indent=2), encoding="utf-8")

    def _read_json(self, path: Path, fallback):
        if not path.exists():
            return fallback
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return fallback

    def _count_by(self, records: list[dict[str, Any]], candidates: list[str]) -> list[dict[str, Any]]:
        header = self._header_lookup_from_records(records, candidates[0], candidates)
        counts: dict[str, int] = {}
        for record in records:
            key = str(record.get(header, "") if header else "").strip() or "Nao informado"
            counts[key] = counts.get(key, 0) + 1
        return [{"label": label, "total": total} for label, total in sorted(counts.items(), key=lambda item: (-item[1], item[0]))[:12]]

    def _count_by_date(self, records: list[dict[str, Any]], candidates: list[str]) -> list[dict[str, Any]]:
        header = self._header_lookup_from_records(records, candidates[0], candidates)
        counts: dict[str, int] = {}
        for record in records:
            key = self._date_label(record.get(header, "") if header else "")
            counts[key] = counts.get(key, 0) + 1
        return [{"label": label, "total": total} for label, total in sorted(counts.items(), key=lambda item: (item[0]))[:12]]

    def _date_label(self, value: Any) -> str:
        if isinstance(value, datetime):
            return value.strftime("%d/%m/%Y")
        text = str(value or "").strip()
        if not text:
            return "Sem data"
        for separator in (" ", "T"):
            if separator in text:
                text = text.split(separator, 1)[0]
        try:
            parsed = datetime.fromisoformat(text)
            return parsed.strftime("%d/%m/%Y")
        except ValueError:
            return text

    def _header_lookup(self, headers: list[str], name: str) -> str | None:
        normalized = self._normalize(name)
        for header in headers:
            if self._normalize(header) == normalized:
                return header
        return None

    def _header_lookup_from_records(self, records: list[dict[str, Any]], name: str, candidates: list[str] | None = None) -> str | None:
        if not records:
            return name
        headers = list(records[0].keys())
        for candidate in candidates or [name]:
            found = self._header_lookup(headers, candidate)
            if found:
                return found
        return None

    def _normalize(self, value: Any) -> str:
        text = str(value or "").strip().upper()
        without_accents = "".join(char for char in unicodedata.normalize("NFD", text) if unicodedata.category(char) != "Mn")
        return "".join(char for char in without_accents if char.isalnum())

    def _is_not_requested(self, value: Any) -> bool:
        return self._normalize(value) in {"NAO", "NO"}

    def _cell_value(self, value: Any) -> Any:
        if isinstance(value, datetime):
            return value.isoformat(sep=" ", timespec="seconds")
        return "" if value is None else value

