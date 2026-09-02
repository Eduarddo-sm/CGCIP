from __future__ import annotations

import copy
import hashlib
from io import BytesIO
from datetime import date, datetime, time
import importlib.util
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class ExcelReadError(Exception):
    pass


class ExcelReader:
    def __init__(self) -> None:
        self._sheet_cache: dict[tuple[str, float, str], list[str]] = {}
        self._table_cache: dict[tuple[str, str, float, str], dict[str, Any]] = {}

    def list_sheets(self, file_path: str, password: str | None = None) -> list[str]:
        path = Path(file_path).expanduser()
        if not path.exists():
            raise ExcelReadError("Arquivo da planilha nao encontrado")
        cache_key = (str(path.resolve()), path.stat().st_mtime, self._password_fingerprint(password))
        if cache_key in self._sheet_cache:
            return list(self._sheet_cache[cache_key])
        if password:
            workbook = self._open_encrypted_workbook(path, password)
            if workbook:
                return self._cache_sheets(cache_key, workbook.sheetnames)
            sheets = self._list_sheets_with_excel(path, password)
            if sheets:
                return self._cache_sheets(cache_key, sheets)
        workbook = self._open_workbook(file_path, password)
        return self._cache_sheets(cache_key, workbook.sheetnames)

    def read_table(self, file_path: str, sheet_name: str, password: str | None = None) -> dict[str, Any]:
        path = Path(file_path).expanduser()
        if not path.exists():
            raise ExcelReadError("Arquivo da planilha nao encontrado")
        cache_key = (str(path.resolve()), sheet_name, path.stat().st_mtime, self._password_fingerprint(password))
        if cache_key in self._table_cache:
            return copy.deepcopy(self._table_cache[cache_key])
        workbook = self._open_workbook(file_path, password)
        if sheet_name not in workbook.sheetnames:
            raise ExcelReadError(f"Sheet '{sheet_name}' nao encontrada")
        sheet = workbook[sheet_name]
        if sheet.protection.sheet:
            sheet.protection.sheet = False
        table_range = self._detect_table_range(sheet)
        rows = list(sheet[table_range])
        if not rows:
            raise ExcelReadError("A sheet selecionada esta vazia")
        headers = [self._stringify(cell.value) or f"Coluna {idx + 1}" for idx, cell in enumerate(rows[0])]
        data_rows = []
        for row_index, row in enumerate(rows[1:], start=2):
            values = [self._normalize(cell.value) for cell in row]
            if all(value in (None, "") for value in values):
                continue
            record = {"_row_id": str(row_index), "_excel_row": row_index}
            for header, value in zip(headers, values):
                record[header] = value
            data_rows.append(record)
        table = {
            "file_path": str(Path(file_path).resolve()),
            "sheet": sheet_name,
            "table_range": table_range,
            "headers": headers,
            "types": self._infer_types(headers, data_rows),
            "rows": data_rows,
            "row_count": len(data_rows),
            "captured_at": datetime.now().isoformat(timespec="seconds"),
        }
        self._table_cache[cache_key] = copy.deepcopy(table)
        return table

    def _cache_sheets(self, cache_key: tuple[str, float, str], sheets: list[str]) -> list[str]:
        self._sheet_cache[cache_key] = list(sheets)
        return list(sheets)

    def _password_fingerprint(self, password: str | None) -> str:
        if not password:
            return ""
        return hashlib.sha256(str(password).encode("utf-8")).hexdigest()

    def _open_workbook(self, file_path: str, password: str | None):
        path = Path(file_path).expanduser()
        if not path.exists():
            raise ExcelReadError("Arquivo da planilha nao encontrado")
        if password:
            workbook = self._open_encrypted_workbook(path, password)
            if workbook:
                return workbook
        try:
            return load_workbook(path, data_only=True, read_only=False)
        except Exception as exc:
            if password:
                if importlib.util.find_spec("msoffcrypto") is None and importlib.util.find_spec("win32com") is None:
                    raise ExcelReadError(
                        "Este arquivo parece exigir senha de abertura. Para buscar sheets com senha, instale "
                        "'msoffcrypto-tool' neste Python ou 'pywin32' para usar o Microsoft Excel instalado."
                    ) from exc
                raise ExcelReadError(
                    "Nao foi possivel abrir o arquivo com a senha informada. "
                    "Confira a senha ou instale 'msoffcrypto-tool' para arquivos Excel criptografados."
                ) from exc
            raise ExcelReadError(f"Nao foi possivel abrir o arquivo Excel: {exc}") from exc

    def _open_encrypted_workbook(self, path: Path, password: str):
        if importlib.util.find_spec("msoffcrypto") is None:
            return None
        import msoffcrypto

        decrypted = BytesIO()
        try:
            with path.open("rb") as source:
                office_file = msoffcrypto.OfficeFile(source)
                if not office_file.is_encrypted():
                    return None
                office_file.load_key(password=password)
                office_file.decrypt(decrypted)
            decrypted.seek(0)
            return load_workbook(decrypted, data_only=True, read_only=False)
        except Exception as exc:
            raise ExcelReadError("Nao foi possivel descriptografar o arquivo Excel com a senha informada.") from exc

    def _list_sheets_with_excel(self, path: Path, password: str) -> list[str] | None:
        if importlib.util.find_spec("win32com") is None or importlib.util.find_spec("pythoncom") is None:
            return None
        import pythoncom
        import win32com.client

        pythoncom.CoInitialize()
        excel = None
        workbook = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            workbook = excel.Workbooks.Open(str(path.resolve()), False, True, None, password)
            return [sheet.Name for sheet in workbook.Worksheets]
        except Exception as exc:
            raise ExcelReadError("Nao foi possivel abrir o arquivo no Excel com a senha informada.") from exc
        finally:
            if workbook is not None:
                workbook.Close(False)
            if excel is not None:
                excel.Quit()
            pythoncom.CoUninitialize()

    def _detect_table_range(self, sheet) -> str:
        tables = list(sheet.tables.values())
        if tables:
            return tables[0].ref
        dimension = sheet.calculate_dimension()
        if dimension == "A1" and sheet["A1"].value is None:
            raise ExcelReadError("Nao ha tabela ou area preenchida na sheet")
        return dimension

    def _infer_types(self, headers: list[str], rows: list[dict[str, Any]]) -> dict[str, str]:
        inferred = {}
        for header in headers:
            values = [row.get(header) for row in rows if row.get(header) not in (None, "")]
            sample = values[:25]
            if not sample:
                inferred[header] = "vazio"
            elif all(isinstance(value, bool) for value in sample):
                inferred[header] = "booleano"
            elif all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in sample):
                inferred[header] = "numero"
            elif all(self._looks_date(value) for value in sample):
                inferred[header] = "data"
            else:
                inferred[header] = "texto"
        return inferred

    def _looks_date(self, value: Any) -> bool:
        return isinstance(value, (datetime, date, time)) or (isinstance(value, str) and len(value) >= 8 and any(sep in value for sep in ("-", "/")))

    def _normalize(self, value: Any) -> Any:
        if isinstance(value, datetime):
            return value.isoformat(timespec="seconds")
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, time):
            return value.isoformat(timespec="seconds")
        return value

    def _stringify(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()
