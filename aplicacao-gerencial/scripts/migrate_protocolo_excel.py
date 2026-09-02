from __future__ import annotations

import json
import sys
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from backend.config import settings  # noqa: E402
from database.repository import Repository  # noqa: E402


CONFIG_PATH = settings.data_dir / "protocolo_config.json"
FIELD_MAP = {
    "DATA": "data_mes",
    "CARTEIRA": "carteira",
    "NOME": "nome",
    "PJ": "pj",
    "PROCESSO": "processo",
    "DATA DE SOLICITACAO": "data_solicitacao",
    "STATUS": "status",
    "DATA DE CONCLUSAO": "data_conclusao",
    "OBSERVACAO": "observacao",
}


def main() -> None:
    config = read_json(CONFIG_PATH)
    excel_path = Path(str(config.get("excel_path") or "")).expanduser()
    if not excel_path.exists():
        raise SystemExit(f"Planilha de protocolo nao encontrada: {excel_path}")
    sheet_name = str(config.get("main_sheet") or "Sol. Protocolo")
    records = read_table(excel_path, sheet_name)
    repo = Repository(settings.database_url)
    imported = repo.upsert_protocolos(records, "migracao_protocolo")
    print(json.dumps({
        "ok": True,
        "excel_path": str(excel_path),
        "sheet": sheet_name,
        "records_read": len(records),
        "records_imported": imported,
    }, ensure_ascii=False))


def read_table(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, keep_vba=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise RuntimeError(f"Aba '{sheet_name}' nao encontrada.")
        sheet = workbook[sheet_name]
        table = find_table(sheet)
        if table is None:
            raise RuntimeError("Nenhuma tabela formatada foi encontrada na aba de protocolo.")
        min_col, header_row, max_col, max_row = range_boundaries(table.ref)
        headers = [
            str(sheet.cell(header_row, col).value or f"COLUNA {col}").strip()
            for col in range(min_col, max_col + 1)
        ]
        result = []
        for row_index in range(header_row + 1, max_row + 1):
            values = {
                headers[offset]: cell_value(sheet.cell(row_index, min_col + offset).value)
                for offset in range(len(headers))
            }
            if not any(value not in ("", None) for value in values.values()):
                continue
            result.append(to_record(row_index, values))
        return result
    finally:
        workbook.close()


def find_table(sheet):
    tables = getattr(sheet, "tables", {})
    best = None
    best_score = -1
    for table in tables.values():
        if not hasattr(table, "ref"):
            table = tables[table]
        min_col, header_row, max_col, _max_row = range_boundaries(table.ref)
        headers = [str(sheet.cell(header_row, col).value or "").strip() for col in range(min_col, max_col + 1)]
        normalized = {normalize(header) for header in headers}
        score = len([header for header in FIELD_MAP if normalize(header) in normalized])
        if score > best_score:
            best = table
            best_score = score
    return best


def to_record(row_key: int, row: dict[str, Any]) -> dict[str, Any]:
    mapped = {"row_key": row_key, "extra": {}, "source": "excel_import"}
    used = set()
    for header, field in FIELD_MAP.items():
        found = find_header(row, header)
        mapped[field] = row.get(found, "") if found else ""
        if found:
            used.add(found)
    mapped["status"] = normalize_status(mapped.get("status"))
    for key, value in row.items():
        if key not in used:
            mapped["extra"][key] = value
    return mapped


def find_header(row: dict[str, Any], name: str) -> str:
    normalized = normalize(name)
    return next((key for key in row if normalize(key) == normalized), "")


def normalize_status(value: Any) -> str:
    return "CONCLUIDO" if normalize(value) == "CONCLUIDO" else "PENDENTE"


def normalize(value: Any) -> str:
    text = str(value or "").strip().upper()
    without_accents = "".join(char for char in unicodedata.normalize("NFD", text) if unicodedata.category(char) != "Mn")
    return "".join(char for char in without_accents if char.isalnum())


def cell_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    return "" if value is None else value


def read_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


if __name__ == "__main__":
    main()
