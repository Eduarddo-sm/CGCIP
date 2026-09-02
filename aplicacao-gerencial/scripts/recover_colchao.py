from __future__ import annotations

import shutil
import traceback
from pathlib import Path

import pythoncom
import win32com.client
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


ROOT = Path(__file__).resolve().parents[1]
LOG = ROOT / "data" / "recovered" / "recover_colchao.log"


def log(message: str) -> None:
    LOG.parent.mkdir(exist_ok=True)
    with LOG.open("a", encoding="utf-8") as file:
        file.write(message + "\n")
    print(message, flush=True)


def main() -> None:
    LOG.parent.mkdir(exist_ok=True)
    LOG.write_text("", encoding="utf-8")
    current = next(Path.home().joinpath("Downloads", "COLCHOES").glob("COLCH*ALPHA.xlsx"))
    healthy = next((ROOT / "data" / "backups").glob("COLCH*145223.xlsx"))
    recovered = ROOT / "data" / "recovered" / "COLCHAO_TOTAL_ALPHA_RECUPERADO.xlsx"
    log(f"CURRENT={current}")
    log(f"HEALTHY={healthy}")
    log(f"RECOVERED={recovered}")
    shutil.copy2(healthy, recovered)

    workbook = load_workbook(current, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    table = next(iter(sheet.tables.values()))
    if not hasattr(table, "ref"):
        table = sheet.tables[table]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = [str(sheet.cell(min_row, col).value).strip() for col in range(min_col, max_col + 1)]
    columns = {}
    for offset, header in enumerate(headers):
        if header == "H.O":
            continue
        columns[header] = [[sheet.cell(row, min_col + offset).value] for row in range(min_row + 1, max_row + 1)]
    workbook.close()
    log(f"COLUMNS={len(columns)} ROWS={len(next(iter(columns.values())))}")

    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        excel_workbook = excel.Workbooks.Open(str(recovered), UpdateLinks=0, ReadOnly=False)
        excel_sheet = excel_workbook.Worksheets(1)
        excel_table = excel_sheet.ListObjects(1)
        excel_headers = {
            str(excel_table.ListColumns(index).Name).strip(): index
            for index in range(1, excel_table.ListColumns.Count + 1)
        }
        log(f"EXCEL_HEADERS={excel_headers}")
        for header, values in columns.items():
            index = excel_headers.get(header)
            log(f"WRITE={header}:{index}")
            if index:
                excel_table.ListColumns(index).DataBodyRange.Value = values
        excel_workbook.Save()
        excel_workbook.Close(SaveChanges=False)
    finally:
        excel.Quit()
        pythoncom.CoUninitialize()

    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        excel_workbook = excel.Workbooks.Open(str(recovered), UpdateLinks=0, ReadOnly=True)
        log("RECOVERED_OK")
        excel_workbook.Close(SaveChanges=False)
    finally:
        excel.Quit()
        pythoncom.CoUninitialize()


if __name__ == "__main__":
    try:
        main()
    except Exception:
        log(traceback.format_exc())
        raise
